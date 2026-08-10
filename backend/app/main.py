from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel


from .config import engine, Base, get_db
from . import models, schemas, crud

# Initialize database schemas
Base.metadata.create_all(bind=engine)

import sqlalchemy
try:
    with engine.connect() as conn:
        conn.execute(sqlalchemy.text("ALTER TABLE daybooks ADD COLUMN is_manually_adjusted INTEGER DEFAULT 0"))
        conn.commit()
except Exception:
    pass

try:
    with engine.connect() as conn:
        try:
            conn.execute(sqlalchemy.text("ALTER TABLE system_users ADD COLUMN current_session_id VARCHAR"))
            conn.commit()
        except Exception:
            pass
        try:
            conn.execute(sqlalchemy.text("ALTER TABLE system_users ADD COLUMN last_active_at TIMESTAMP"))
            conn.commit()
        except Exception:
            pass
except Exception:
    pass

# ─── SECURITY & AUTHENTICATION UTILITIES ───
import hashlib
import secrets
import base64
import hmac
import json
import time

SECRET_KEY = "pooja-jewellers-super-secret-key-that-is-extremely-secure"
GLOBAL_TIME_OFFSET_SECONDS = 0.0

def get_real_ist_now():
    from datetime import datetime, timezone, timedelta
    ist_tz = timezone(timedelta(hours=5, minutes=30))
    now_ts = time.time() + GLOBAL_TIME_OFFSET_SECONDS
    return datetime.fromtimestamp(now_ts, tz=ist_tz)


def hash_password(password: str, salt: bytes = None) -> str:
    if salt is None:
        salt = secrets.token_bytes(16)
    kdf = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 100000)
    return f"{salt.hex()}:{kdf.hex()}"

def verify_password(password: str, hashed: str) -> bool:
    try:
        salt_str, key_str = hashed.split(":")
        salt = bytes.fromhex(salt_str)
        key = bytes.fromhex(key_str)
        kdf = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 100000)
        return secrets.compare_digest(kdf, key)
    except Exception:
        return False

def base64url_encode(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode("utf-8")

def base64url_decode(data: str) -> bytes:
    padding = "=" * (4 - (len(data) % 4))
    return base64.urlsafe_b64decode(data + padding)

def create_access_token(subject: str, session_id: Optional[str] = None, expires_in: int = 86400) -> str:
    header = {"alg": "HS256", "typ": "JWT"}
    payload = {
        "sub": subject,
        "exp": int(time.time()) + expires_in
    }
    if session_id:
        payload["sid"] = session_id
    header_b64 = base64url_encode(json.dumps(header).encode("utf-8"))
    payload_b64 = base64url_encode(json.dumps(payload).encode("utf-8"))
    message = f"{header_b64}.{payload_b64}"
    signature = hmac.new(SECRET_KEY.encode("utf-8"), message.encode("utf-8"), hashlib.sha256).digest()
    signature_b64 = base64url_encode(signature)
    return f"{message}.{signature_b64}"

def decode_access_token(token: str) -> dict:
    try:
        parts = token.split(".")
        if len(parts) != 3:
            return None
        header_b64, payload_b64, signature_b64 = parts
        message = f"{header_b64}.{payload_b64}"
        expected_sig = hmac.new(SECRET_KEY.encode("utf-8"), message.encode("utf-8"), hashlib.sha256).digest()
        expected_sig_b64 = base64url_encode(expected_sig)
        if not hmac.compare_digest(signature_b64, expected_sig_b64):
            return None
        payload = json.loads(base64url_decode(payload_b64).decode("utf-8"))
        if payload.get("exp", 0) < time.time():
            return None
        return payload
    except Exception:
        return None

# Auto-seed default user "admin" with password "pooja" if not already present
def seed_default_admin():
    from sqlalchemy.orm import sessionmaker
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    db = SessionLocal()
    try:
        admin_user = db.query(models.SystemUser).filter(models.SystemUser.username == "admin").first()
        if not admin_user:
            default_pw = "pooja"
            hashed = hash_password(default_pw)
            db.add(models.SystemUser(username="admin", password_hash=hashed))
            db.commit()
            print("[Startup] Seeded default admin user successfully.")
    except Exception as e:
        db.rollback()
        print("[Startup] Failed to seed default admin user:", e)
    finally:
        db.close()

seed_default_admin()

from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi import Request

security_scheme = HTTPBearer(auto_error=False)

def verify_token_dependency(
    request: Request, 
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security_scheme),
    db: Session = Depends(get_db)
):
    path = request.url.path
    # Bypassed paths
    if path in ["/api/auth/login", "/api/system/run-id", "/api/system/network-time", "/api/system/google-time", "/api/system/open-date-settings", "/api/system/set-time-offset", "/api/setup/is-first-time", "/api/live-rates", "/api/udar/outstanding"] or path.startswith("/api/whatsapp") or path.startswith("/api/udar"):
        return None
        
    if not credentials:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Not authenticated. Missing Authorization credentials."
        )
        
    token = credentials.credentials
    payload = decode_access_token(token)
    if not payload:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Session expired or invalid token. Please log in again."
        )
        
    username = payload.get("sub")
    sid = payload.get("sid")
    if not sid:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="SESSION_SUPERSEDED"
        )
        
    if username:
        user = db.query(models.SystemUser).filter(models.SystemUser.username == username).first()
        if user:
            if user.current_session_id != sid:
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail="SESSION_SUPERSEDED"
                )
            from datetime import datetime
            user.last_active_at = datetime.utcnow()
            try:
                db.commit()
            except Exception:
                db.rollback()
                
    return payload


# Auto-migrations for PledgeEntry new columns
def run_migrations():
    import sqlite3
    import os
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, "..", "database.db")
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(pledge_entries)")
        columns = [col[1] for col in cursor.fetchall()]
        
        new_cols = [
            ("pledge_no", "TEXT"),
            ("pawner_relation", "TEXT"),
            ("pawner_relation_name", "TEXT"),
            ("mobile", "TEXT"),
            ("income", "TEXT"),
            ("address", "TEXT"),
            ("rupees_in_words", "TEXT"),
            ("interest_rate_text", "TEXT"),
            ("redemption_period_months", "INTEGER"),
            ("interest_payment_frequency", "TEXT"),
            ("gross_weight", "REAL"),
            ("less_weight", "REAL"),
            ("net_weight", "REAL"),
            ("quantity", "INTEGER"),
            ("estimated_value", "REAL"),
            ("due_date", "TEXT"),
            ("status", "TEXT DEFAULT 'ACTIVE'"),
            ("release_date", "TEXT"),
            ("customer_photo", "TEXT"),
            ("item_photo", "TEXT"),
            ("ornament_2", "TEXT"),
            ("quantity_2", "INTEGER"),
            ("gross_weight_2", "REAL"),
            ("less_weight_2", "REAL"),
            ("net_weight_2", "REAL"),
            ("estimated_value_2", "REAL"),
            ("ornament_3", "TEXT"),
            ("quantity_3", "INTEGER"),
            ("gross_weight_3", "REAL"),
            ("less_weight_3", "REAL"),
            ("net_weight_3", "REAL"),
            ("estimated_value_3", "REAL"),
            ("is_existing", "INTEGER DEFAULT 0"),
            ("is_repledged", "INTEGER DEFAULT 0"),
            ("repledged_bank", "TEXT"),
            ("repledged_amount", "REAL"),
            ("repledged_date", "TEXT"),
            ("repledged_name", "TEXT"),
            ("repledged_receipt_no", "TEXT"),
            ("repledged_entries", "TEXT"),
            ("repledged_interest_amount", "REAL"),
            ("repledged_interest_rate", "TEXT")
        ]
        
        for col_name, col_type in new_cols:
            if col_name not in columns:
                cursor.execute(f"ALTER TABLE pledge_entries ADD COLUMN {col_name} {col_type}")
                print(f"[Migration] Added column {col_name} to pledge_entries")
        
        # Daybooks migrations
        cursor.execute("PRAGMA table_info(daybooks)")
        daybook_columns = [col[1] for col in cursor.fetchall()]
        daybook_new_cols = [
            ("opening_upi_details", "TEXT DEFAULT '{}'"),
            ("closing_upi_details", "TEXT DEFAULT '{}'")
        ]
        for col_name, col_type in daybook_new_cols:
            if col_name not in daybook_columns:
                cursor.execute(f"ALTER TABLE daybooks ADD COLUMN {col_name} {col_type}")
                print(f"[Migration] Added column {col_name} to daybooks")

        # PurchaseItem migrations
        cursor.execute("PRAGMA table_info(purchase_items)")
        pi_cols = [col[1] for col in cursor.fetchall()]
        pi_new_cols = [
            ("is_sold", "INTEGER DEFAULT 0"),
            ("sold_date", "TEXT"),
        ]
        for col_name, col_type in pi_new_cols:
            if col_name not in pi_cols:
                cursor.execute(f"ALTER TABLE purchase_items ADD COLUMN {col_name} {col_type}")
                print(f"[Migration] Added column {col_name} to purchase_items")

        # SoldExcelBarcode migrations
        cursor.execute("PRAGMA table_info(sold_excel_barcodes)")
        seb_cols = [col[1] for col in cursor.fetchall()]
        if "sold_date" not in seb_cols:
            cursor.execute("ALTER TABLE sold_excel_barcodes ADD COLUMN sold_date TEXT")
            print("[Migration] Added column sold_date to sold_excel_barcodes")

        # PurchaseBill migrations
        cursor.execute("PRAGMA table_info(purchase_bills)")
        pb_cols = [col[1] for col in cursor.fetchall()]
        pb_new_cols = [
            ("total_weight", "REAL"),
            ("purity", "TEXT"),
            ("is_rate_cut", "INTEGER DEFAULT 1"),
            ("rate", "REAL"),
            ("amount", "REAL"),
            ("pure_weight", "REAL")
        ]
        for col_name, col_type in pb_new_cols:
            if col_name not in pb_cols:
                cursor.execute(f"ALTER TABLE purchase_bills ADD COLUMN {col_name} {col_type}")
                print(f"[Migration] Added column {col_name} to purchase_bills")

        # SystemLog migrations
        cursor.execute("PRAGMA table_info(system_logs)")
        slog_cols = [col[1] for col in cursor.fetchall()]
        slog_new_cols = [
            ("module", "TEXT DEFAULT 'GENERAL'"),
            ("user_name", "TEXT DEFAULT 'admin'")
        ]
        for col_name, col_type in slog_new_cols:
            if col_name not in slog_cols:
                cursor.execute(f"ALTER TABLE system_logs ADD COLUMN {col_name} {col_type}")
                print(f"[Migration] Added column {col_name} to system_logs")

        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[Migration] Error during table migration: {e}")

run_migrations()

# Daily Automated Backup System
def start_backup_scheduler():
    import threading
    import shutil
    import datetime
    import time
    import os

    def backup_loop():
        # Delay initial run slightly to not slow down start
        time.sleep(15)
        while True:
            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                db_path = os.path.join(base_dir, "..", "database.db")
                backups_dir = os.path.join(base_dir, "..", "backups")
                if not os.path.exists(backups_dir):
                    os.makedirs(backups_dir)
                
                if os.path.exists(db_path):
                    today_str = datetime.datetime.now().strftime("%Y-%m-%d")
                    backup_filename = f"db_backup_{today_str}.db"
                    backup_file_path = os.path.join(backups_dir, backup_filename)
                    
                    shutil.copy2(db_path, backup_file_path)
                    print(f"[Backup Scheduler] Created daily backup: {backup_filename}")
                    
                    backup_files = [
                        os.path.join(backups_dir, f) for f in os.listdir(backups_dir)
                        if f.startswith("db_backup_") and f.endswith(".db")
                    ]
                    backup_files.sort(key=os.path.getmtime)
                    if len(backup_files) > 14:
                        to_remove = backup_files[:-14]
                        for f in to_remove:
                            try:
                                os.remove(f)
                                print(f"[Backup Scheduler] Rotated out old backup: {os.path.basename(f)}")
                            except Exception as ree:
                                print(f"[Backup Scheduler] Failed to delete old backup {f}: {ree}")
            except Exception as e:
                print(f"[Backup Scheduler] Error during backup: {e}")
            # Run every 24 hours
            time.sleep(86400)

    t = threading.Thread(target=backup_loop, daemon=True)
    t.start()

start_backup_scheduler()

app = FastAPI(
    title="Marwadi Digital Day Book API",
    description="Backend API for Pooja Jewellers Day Book diary",
    version="1.0.0",
    dependencies=[Depends(verify_token_dependency)]
)

# Enable CORS for Next.js frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "https://day-book-five.vercel.app"
    ],
    allow_origin_regex=r"https://.*\.vercel\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

import uuid
SERVER_RUN_ID = str(uuid.uuid4())

def log_system_action(db: Session, action: str, details: str):
    try:
        import datetime
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        db_log = models.SystemLog(timestamp=now_str, action=action, details=details)
        db.add(db_log)
        db.commit()
    except Exception as e:
        print(f"[Error Logging Action] {action}: {details}. Error: {e}")

@app.get("/api/system/run-id")
def get_run_id():
    return {"run_id": SERVER_RUN_ID}

@app.get("/api/system/google-time")
def get_google_time():
    """Check both system time and Google/network time to detect any time alterations."""
    import urllib.request
    import email.utils
    import json
    from datetime import datetime, timezone

    google_timestamp = None
    source = "google"

    # 1. Fetch Google HTTP Date header
    try:
        req = urllib.request.Request("https://www.google.com", headers={"User-Agent": "Mozilla/5.0"}, method="HEAD")
        with urllib.request.urlopen(req, timeout=3) as resp:
            date_header = resp.headers.get("Date")
            if date_header:
                parsed_tuple = email.utils.parsedate_tz(date_header)
                if parsed_tuple:
                    dt = email.utils.mktime_tz(parsed_tuple)
                    google_timestamp = int(dt * 1000)
    except Exception as e:
        print("[Time Check] Google time fetch failed:", e)

    # 2. Fallback to timeapi.io if Google HTTP header was unreachable
    if not google_timestamp:
        try:
            req = urllib.request.Request("https://timeapi.io/api/v1/time/current/zone?timeZone=Asia/Kolkata", headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=3) as resp:
                data = json.loads(resp.read().decode())
                if "date_time" in data:
                    source = "timeapi.io"
                    dt = datetime.fromisoformat(data["date_time"])
                    google_timestamp = int(dt.timestamp() * 1000)
        except Exception as e:
            print("[Time Check] timeapi.io fetch failed:", e)

    # 3. Fallback to WorldTimeAPI if Google and timeapi.io were unreachable
    if not google_timestamp:
        try:
            req = urllib.request.Request("https://worldtimeapi.org/api/timezone/Asia/Kolkata", headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=3) as resp:
                data = json.loads(resp.read().decode())
                if "datetime" in data:
                    source = "worldtimeapi"
                    dt = datetime.fromisoformat(data["datetime"].replace("Z", "+00:00"))
                    google_timestamp = int(dt.timestamp() * 1000)
        except Exception as e:
            print("[Time Check] WorldTimeAPI fetch failed:", e)

    system_timestamp = int(datetime.now(timezone.utc).timestamp() * 1000)

    if google_timestamp:
        diff_seconds = abs(google_timestamp - system_timestamp) / 1000.0
        mismatch = diff_seconds > 60  # > 1 minute difference
        return {
            "success": True,
            "google_timestamp": google_timestamp,
            "system_timestamp": system_timestamp,
            "diff_seconds": diff_seconds,
            "diff_minutes": round(diff_seconds / 60.0, 1),
            "source": source,
            "mismatch": mismatch
        }
    else:
        return {
            "success": False,
            "google_timestamp": system_timestamp,
            "system_timestamp": system_timestamp,
            "diff_seconds": 0,
            "diff_minutes": 0,
            "source": "fallback_system",
            "mismatch": False
        }

@app.post("/api/system/open-date-settings")
def open_date_settings():
    """Launch Windows Date & Time Settings directly on the device."""
    import os, platform, subprocess
    try:
        if platform.system() == "Windows":
            os.system("start ms-settings:dateandtime")
            return {"success": True, "message": "Opened Windows Date & Time Settings"}
        else:
            return {"success": False, "message": "Only supported on Windows"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/api/auth/login")
def login(payload: dict, db: Session = Depends(get_db)):
    username = payload.get("username", "admin") or "admin"
    password = payload.get("password")
    force = payload.get("force", False)
    
    if not password:
        raise HTTPException(status_code=400, detail="Password is required")
        
    user = db.query(models.SystemUser).filter(models.SystemUser.username == username).first()
    if not user:
        # Fallback to the main admin user so any username (like "pooja") works
        user = db.query(models.SystemUser).filter(models.SystemUser.username == "admin").first()
        
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Incorrect username or password"
        )
        
    is_valid = verify_password(password, user.password_hash)
    if not is_valid:
        # Backward compatibility: support "pooja123" if the stored password is still the default "pooja"
        if user.username == "admin" and password == "pooja123":
            if verify_password("pooja", user.password_hash):
                is_valid = True
                
    if not is_valid:
        raise HTTPException(
            status_code=401,
            detail="Incorrect username or password"
        )
        
    # Check if there is already an active session in the last 2 minutes
    from datetime import datetime, timedelta
    import uuid
    
    now = datetime.utcnow()
    is_session_active = False
    if user.current_session_id and user.last_active_at:
        last_active = user.last_active_at
        if isinstance(last_active, str):
            try:
                clean_str = last_active.split(".")[0]
                last_active = datetime.strptime(clean_str, "%Y-%m-%d %H:%M:%S")
            except Exception:
                last_active = now - timedelta(minutes=5)
        try:
            if (now - last_active) < timedelta(minutes=2):
                is_session_active = True
        except Exception:
            pass
            
    if is_session_active and not force:
        return {"status": "session_active", "message": "Another device is currently logged in."}
        
    # Generate new session ID and save it
    new_sid = str(uuid.uuid4())
    user.current_session_id = new_sid
    user.last_active_at = now
    db.commit()
        
    token = create_access_token(user.username, session_id=new_sid)
    return {"access_token": token, "token_type": "bearer"}

# --- DayBook Routes ---


@app.get("/api/setup/is-first-time")
def is_first_time_setup(db: Session = Depends(get_db)):
    """Returns true if no daybook has ever had an opening balance set and there are no transactions."""
    from sqlalchemy import func
    total = db.query(func.count(models.DayBook.id)).scalar()
    if total == 0:
        return {"first_time": True}
    
    # Check if any transaction exists
    has_debits = db.query(models.DebitEntry).first() is not None
    has_credits = db.query(models.CreditEntry).first() is not None
    has_sold = db.query(models.SoldItem).first() is not None
    has_phonepe = db.query(models.PhonePeEntry).first() is not None
    has_pledges = db.query(models.PledgeEntry).first() is not None
    has_releases = db.query(models.ReleaseEntry).first() is not None

    if has_debits or has_credits or has_sold or has_phonepe or has_pledges or has_releases:
        return {"first_time": False}

    # Check if any daybook has a non-zero opening balance
    has_opening = db.query(models.DayBook).filter(
        (models.DayBook.opening_cash > 0) |
        (models.DayBook.opening_upi > 0) |
        (models.DayBook.opening_other > 0)
    ).first()
    return {"first_time": has_opening is None}
@app.get("/api/dashboard/stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    # Get last 7 daybooks
    recent = db.query(models.DayBook).order_by(models.DayBook.date.desc()).limit(7).all()
    # Reverse to show chronologically
    recent.reverse()
    
    # Calculate lifetime stats
    from sqlalchemy import func
    import re
    total_days = db.query(func.count(models.DayBook.id)).scalar()
    
    # Gold and Silver sold
    sold_items = db.query(models.SoldItem).all()
    gold_sold = sum(item.weight for item in sold_items if "[GOLD]" in item.item_name)
    silver_sold = sum(item.weight for item in sold_items if "[SILVER]" in item.item_name)
    
    # Pledges outstanding (Pledge entries that are NOT released)
    total_pledged = db.query(func.sum(models.PledgeEntry.amount)).scalar() or 0.0
    total_released = db.query(func.sum(models.ReleaseEntry.principal_amount)).scalar() or 0.0
    outstanding_girvi = max(0.0, total_pledged - total_released)

    # ─── CALCULATE STOCK WEIGHTS & VALUATION ───
    try:
        load_barcode_cache()
    except Exception:
        pass

    # Build set of sold barcodes
    sold_barcodes = set()
    for item in sold_items:
        if item.item_name:
            match = re.search(r"\[BARCODE:([A-Za-z0-9]+)\]", item.item_name)
            if match:
                sold_barcodes.add(match.group(1).strip().upper())
                
    manual_sold = db.query(models.SoldExcelBarcode.barcode_no).all()
    for (bc,) in manual_sold:
        if bc:
            sold_barcodes.add(bc.strip().upper())

    # Excel stock weights (available only)
    excel_gold_wt = 0.0
    excel_silver_wt = 0.0
    for bc, details in BARCODE_CACHE.items():
        bc_upper = bc.strip().upper()
        if bc_upper in sold_barcodes:
            continue
        wt = details.get("weight", 0.0)
        metal = details.get("metal", "GOLD")
        if metal == "GOLD":
            excel_gold_wt += wt
        elif metal == "SILVER":
            excel_silver_wt += wt

    # DB stock weights (available only)
    db_gold_wt = 0.0
    db_silver_wt = 0.0
    db_items = db.query(models.PurchaseItem).all()
    for pi in db_items:
        bc_upper = pi.barcode_no.strip().upper()
        is_sold = 1 if (pi.is_sold == 1 or bc_upper in sold_barcodes) else 0
        if is_sold:
            continue
        wt = pi.weight or 0.0
        metal = pi.metal or "GOLD"
        if metal == "GOLD":
            db_gold_wt += wt
        elif metal == "SILVER":
            db_silver_wt += wt

    total_gold_stock = excel_gold_wt + db_gold_wt
    total_silver_stock = excel_silver_wt + db_silver_wt

    # Scrape or use cache live rate
    def parse_rate(rate_str: str) -> float:
        try:
            return float(rate_str.replace(",", "").strip())
        except Exception:
            return 0.0

    gold_rate_val = parse_rate(LIVE_RATES_CACHE.get("gold_22k", "0"))
    if gold_rate_val <= 0:
        gold_rate_val = 7200.0  # fallback: 22K gold rate
        
    silver_rate_val = parse_rate(LIVE_RATES_CACHE.get("silver", "0"))
    if silver_rate_val <= 0:
        silver_rate_val = 90000.0  # fallback per kg
    silver_per_gram = silver_rate_val / 1000.0

    stock_val_gold = total_gold_stock * gold_rate_val
    stock_val_silver = total_silver_stock * silver_per_gram
    total_stock_valuation = stock_val_gold + stock_val_silver

    # ─── CALCULATE TOTAL SUPPLIER CREDIT DYNAMICALLY ───
    registered_parties = db.query(models.PurchaseParty).all()
    bill_suppliers = db.query(models.PurchaseBill.supplier_name).distinct().all()
    
    supplier_names = set(p.name.strip().lower() for p in registered_parties)
    for (name,) in bill_suppliers:
        if name and name.strip():
            supplier_names.add(name.strip().lower())
            
    total_supplier_credit = 0.0
    for name_lower in supplier_names:
        # Sum bills
        bills_sum = db.query(func.sum(models.PurchaseBill.invoice_total)).filter(
            func.lower(models.PurchaseBill.supplier_name) == name_lower
        ).scalar() or 0.0
        
        # Sum cash payments
        debits_sum = db.query(func.sum(models.DebitEntry.amount)).filter(
            func.lower(models.DebitEntry.name) == name_lower,
            ~models.DebitEntry.particulars.like("%[RATE_CUT:%")
        ).scalar() or 0.0
        
        # Get opening balance if registered
        party_rec = next((p for p in registered_parties if p.name.strip().lower() == name_lower), None)
        opening_cash = party_rec.opening_balance_cash if party_rec else 0.0
        
        outstanding_cash = opening_cash + bills_sum - debits_sum
        total_supplier_credit += outstanding_cash

    # ─── CALCULATE ACTIVE PLEDGE DISTRIBUTION ───
    active_pledges = db.query(models.PledgeEntry).filter(models.PledgeEntry.status == "ACTIVE").all()
    active_gold_val = 0.0
    active_silver_val = 0.0
    active_gold_wt = 0.0
    active_silver_wt = 0.0
    
    for p in active_pledges:
        is_silver_1 = any(x in (p.ornament or "").lower() for x in ["silver", "chandi", "sil"])
        w1 = p.net_weight or p.weight or 0.0
        w2 = p.net_weight_2 or 0.0
        w3 = p.net_weight_3 or 0.0
        
        if is_silver_1:
            active_silver_wt += w1
            active_silver_val += p.amount
        else:
            active_gold_wt += w1
            active_gold_val += p.amount
            
        if p.ornament_2:
            is_silver_2 = any(x in p.ornament_2.lower() for x in ["silver", "chandi", "sil"])
            if is_silver_2:
                active_silver_wt += w2
            else:
                active_gold_wt += w2
                
        if p.ornament_3:
            is_silver_3 = any(x in p.ornament_3.lower() for x in ["silver", "chandi", "sil"])
            if is_silver_3:
                active_silver_wt += w3
            else:
                active_gold_wt += w3

    return {
        "recent_days": [
            {
                "date": d.date,
                "closing_cash": d.closing_cash,
                "closing_upi": d.closing_upi,
                "closing_other": d.closing_other,
                "total": d.closing_cash + d.closing_upi + d.closing_other
            } for d in recent
        ],
        "total_days": total_days,
        "gold_sold": gold_sold,
        "silver_sold": silver_sold,
        "outstanding_girvi": outstanding_girvi,
        "total_gold_stock": total_gold_stock,
        "total_silver_stock": total_silver_stock,
        "total_stock_valuation": total_stock_valuation,
        "total_supplier_credit": total_supplier_credit,
        "active_gold_val": active_gold_val,
        "active_silver_val": active_silver_val,
        "active_gold_wt": active_gold_wt,
        "active_silver_wt": active_silver_wt
    }

@app.get("/api/reports/range", response_model=List[schemas.DayBookResponse])
def get_reports_range(start_date: str, end_date: str, db: Session = Depends(get_db)):
    daybooks = db.query(models.DayBook).filter(
        models.DayBook.date >= start_date,
        models.DayBook.date <= end_date
    ).order_by(models.DayBook.date.asc()).all()
    return daybooks

@app.get("/api/reports/aavak-jaavak")
def get_aavak_jaavak_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    flow_type: Optional[str] = None,
    payment_mode: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    udhar_filter: Optional[str] = None,
    db: Session = Depends(get_db)
):
    return crud.get_master_aavak_jaavak_entries(
        db,
        start_date=start_date,
        end_date=end_date,
        flow_type=flow_type,
        payment_mode=payment_mode,
        category=category,
        search=search,
        udhar_filter=udhar_filter
    )



class WhatsAppPDFRequest(BaseModel):
    invoice_no: str
    book_no: Optional[str] = ""
    date: str
    customer_name: str
    phone: Optional[str] = ""
    address: Optional[str] = ""
    items: List[dict]
    total_amount: float
    message: Optional[str] = ""

@app.post("/api/whatsapp/send-pdf-bill")
def api_send_whatsapp_pdf_bill(req: WhatsAppPDFRequest):
    from app.whatsapp_service import create_pdf_invoice_file, sanitize_phone_number, dispatch_whatsapp_bill_directly
    pdf_path = create_pdf_invoice_file(
        invoice_no=req.invoice_no,
        date_str=req.date,
        customer_name=req.customer_name,
        items=req.items,
        total_amount=req.total_amount,
        book_no=req.book_no or "",
        phone=req.phone or "",
        address=req.address or ""
    )
    
    filename = os.path.basename(pdf_path)
    clean_phone = sanitize_phone_number(req.phone or "")
    dispatch_whatsapp_bill_directly(req.phone or "", pdf_path, req.message or "")
    
    return {
        "status": "success",
        "message": "PDF Invoice generated and dispatched successfully",
        "file_name": filename,
        "download_url": f"/api/whatsapp/download-pdf/{filename}",
        "phone": clean_phone
    }


@app.get("/api/whatsapp/download-pdf/{filename}")
def api_download_whatsapp_pdf(filename: str):
    from app.whatsapp_service import INVOICES_DIR
    from fastapi.responses import FileResponse
    safe_filename = os.path.basename(filename)
    file_path = os.path.join(INVOICES_DIR, safe_filename)
    if os.path.exists(file_path):
        return FileResponse(file_path, media_type="application/pdf", filename=safe_filename)
    raise HTTPException(status_code=404, detail="PDF invoice file not found")



@app.get("/api/udar/outstanding")
def get_outstanding_udar(db: Session = Depends(get_db)):
    debits = db.query(models.DebitEntry).all()
    credits = db.query(models.CreditEntry).all()
    
    debit_totals = {}
    for d in debits:
        part = d.particulars.lower()
        rem = (d.remarks or "").lower()
        if "udar" in part or "udhar" in part or "udar" in rem or "udhar" in rem:
            name_key = d.name.strip()
            debit_totals[name_key] = debit_totals.get(name_key, 0.0) + d.amount
            
    credit_totals = {}
    for c in credits:
        part = c.particulars.lower()
        rem = (c.remarks or "").lower()
        if "udar" in part or "udhar" in part or "return" in part or "udar" in rem or "udhar" in rem or "return" in rem:
            name_key = c.name.strip()
            credit_totals[name_key] = credit_totals.get(name_key, 0.0) + c.amount
            
    outstanding = []
    name_map = {}
    
    for name, debit_amt in debit_totals.items():
        name_lower = name.lower()
        if name_lower not in name_map:
            name_map[name_lower] = {"original_name": name, "debit": 0.0, "credit": 0.0}
        name_map[name_lower]["debit"] += debit_amt
        
    for name, credit_amt in credit_totals.items():
        name_lower = name.lower()
        if name_lower not in name_map:
            name_map[name_lower] = {"original_name": name, "debit": 0.0, "credit": 0.0}
        name_map[name_lower]["credit"] += credit_amt
        
    for name_lower, data in name_map.items():
        net = data["debit"] - data["credit"]
        if net != 0.0:
            outstanding.append({
                "name": data["original_name"],
                "amount": net
            })
            
    return outstanding


@app.get("/api/daybook/date/{date_str}", response_model=schemas.DayBookResponse)
def get_daybook_by_date(date_str: str, db: Session = Depends(get_db)):
    prev = crud.get_previous_daybook(db, date_str)
    if prev:
        crud.recalculate_and_cascade_daybook(db, prev.id)
        db.refresh(prev)

    p_cash = (prev.closing_cash or 0.0) if prev else 0.0
    p_upi = (prev.closing_upi or 0.0) if prev else 0.0
    p_other = (prev.closing_other or 0.0) if prev else 0.0
    p_upi_details = (prev.closing_upi_details or "{}") if prev else "{}"

    db_daybook = crud.get_daybook_by_date(db, date_str)

    if db_daybook:
        if prev and getattr(db_daybook, "is_manually_adjusted", 0) != 1:
            db_daybook.opening_cash = p_cash
            db_daybook.opening_upi = p_upi
            db_daybook.opening_other = p_other
            db_daybook.opening_upi_details = p_upi_details
            db.commit()
        crud.recalculate_and_cascade_daybook(db, db_daybook.id)
        db.refresh(db_daybook)
        return db_daybook

    # ── Brand-new day: return virtual in-memory response WITHOUT saving to DB ──
    return schemas.DayBookResponse(
        id=-1,
        date=date_str,
        opening_cash=p_cash,
        opening_upi=p_upi,
        opening_other=p_other,
        closing_cash=p_cash,
        closing_upi=p_upi,
        closing_other=p_other,
        opening_upi_details=p_upi_details,
        closing_upi_details=p_upi_details,
        is_manually_adjusted=0,
        debit_entries=[],
        credit_entries=[],
        sold_items=[],
        phonepe_entries=[],
        old_gold_entries=[],
        old_silver_entries=[],
        pledge_entries=[],
        release_entries=[],
        pledge_payments=[]
    )

@app.post("/api/daybook", response_model=schemas.DayBookResponse)
def create_daybook(daybook: schemas.DayBookCreate, db: Session = Depends(get_db)):
    existing = crud.get_daybook_by_date(db, daybook.date)
    if existing:
        return existing
    res = crud.create_daybook(db, daybook)
    log_system_action(db, "DAYBOOK_CREATE", f"Created DayBook for date {daybook.date} with opening cash ₹{daybook.opening_cash}, UPI ₹{daybook.opening_upi}")
    return res

@app.put("/api/daybook/{daybook_id}", response_model=schemas.DayBookResponse)
def update_daybook(daybook_id: int, daybook_update: schemas.DayBookUpdate, db: Session = Depends(get_db)):
    updated = crud.update_daybook_cash(
        db, daybook_id,
        opening_cash=daybook_update.opening_cash,
        opening_upi=daybook_update.opening_upi,
        opening_other=daybook_update.opening_other,
        closing_cash=daybook_update.closing_cash,
        closing_upi=daybook_update.closing_upi,
        closing_other=daybook_update.closing_other,
        opening_upi_details=daybook_update.opening_upi_details,
        closing_upi_details=daybook_update.closing_upi_details
    )
    if not updated:
        raise HTTPException(status_code=404, detail="DayBook not found")
    log_system_action(db, "DAYBOOK_UPDATE", f"Updated DayBook settings for {updated.date} (opening: cash ₹{updated.opening_cash}, UPI ₹{updated.opening_upi})")
    return updated

@app.delete("/api/daybook/{daybook_id}")
def delete_daybook(daybook_id: int, db: Session = Depends(get_db)):
    db_day = db.query(models.DayBook).filter(models.DayBook.id == daybook_id).first()
    date_str = db_day.date if db_day else str(daybook_id)
    success = crud.delete_daybook(db, daybook_id)
    if not success:
        raise HTTPException(status_code=404, detail="DayBook not found")
    log_system_action(db, "DAYBOOK_DELETE", f"Deleted DayBook for date {date_str}")
    return {"message": "DayBook deleted successfully"}



# --- Sub-Entry REST APIs ---

# DEBIT
@app.post("/api/daybook/{daybook_id}/debit", response_model=schemas.DebitEntry)
def add_debit_entry(daybook_id: int, entry: schemas.DebitEntryCreate, db: Session = Depends(get_db)):
    daybook = crud.get_daybook(db, daybook_id)
    if not daybook:
        raise HTTPException(status_code=404, detail="DayBook not found")
    res = crud.create_debit_entry(db, entry, daybook_id)
    log_system_action(db, "DEBIT_CREATE", f"[{daybook.date}] Added Debit: {entry.name} | {entry.particulars} | ₹{entry.amount}")
    return res

@app.put("/api/debit/{entry_id}", response_model=schemas.DebitEntry)
def update_debit_entry(entry_id: int, entry_update: schemas.DebitEntryUpdate, db: Session = Depends(get_db)):
    res = crud.update_debit_entry(db, entry_id, entry_update)
    if not res:
        raise HTTPException(status_code=404, detail="Debit entry not found")
    log_system_action(db, "DEBIT_UPDATE", f"Updated Debit #{entry_id}: {res.name} | {res.particulars} | ₹{res.amount}")
    return res

@app.delete("/api/debit/{entry_id}")
def delete_debit_entry(entry_id: int, db: Session = Depends(get_db)):
    db_entry = db.query(models.DebitEntry).filter(models.DebitEntry.id == entry_id).first()
    if not db_entry:
        raise HTTPException(status_code=404, detail="Debit entry not found")
    details = f"Deleted Debit: {db_entry.name} | {db_entry.particulars} | ₹{db_entry.amount}"
    if not crud.delete_debit_entry(db, entry_id):
        raise HTTPException(status_code=404, detail="Debit entry not found")
    log_system_action(db, "DEBIT_DELETE", details)
    return {"message": "Debit entry deleted"}

class ConvertRateCutRequest(BaseModel):
    rate: float
    metal: str  # "GOLD" or "SILVER"

@app.post("/api/debit/{entry_id}/convert-rate-cut")
def convert_debit_to_rate_cut(entry_id: int, body: ConvertRateCutRequest, db: Session = Depends(get_db)):
    """Convert an existing debit payment into a metal Rate Cut entry."""
    import re
    db_debit = db.query(models.DebitEntry).filter(models.DebitEntry.id == entry_id).first()
    if not db_debit:
        raise HTTPException(status_code=404, detail="Payment entry not found")
        
    if body.rate <= 0:
        raise HTTPException(status_code=400, detail="Rate must be greater than 0")
        
    metal_upper = body.metal.upper()
    if metal_upper not in ["GOLD", "SILVER"]:
        raise HTTPException(status_code=400, detail="Metal must be GOLD or SILVER")
        
    # Strip any existing [RATE_CUT:...] tag if already present
    clean_particulars = re.sub(r"\[RATE_CUT:[^\]]+\]\s*", "", db_debit.particulars or "")
    
    weight = round(db_debit.amount / body.rate, 3)
    rc_tag = f"[RATE_CUT:{metal_upper}|{body.rate}|{weight}] "
    
    db_debit.particulars = rc_tag + clean_particulars.strip()
    db.commit()
    db.refresh(db_debit)
    
    if db_debit.daybook_id:
        crud.recalculate_and_cascade_daybook(db, db_debit.daybook_id)
        
    log_system_action(db, "CONVERT_RATE_CUT", f"Converted Debit #{entry_id} (₹{db_debit.amount}) to Rate Cut: {weight}g {metal_upper} @ ₹{body.rate}/g for supplier '{db_debit.name}'")
    return {
        "success": True,
        "entry_id": db_debit.id,
        "particulars": db_debit.particulars,
        "metal": metal_upper,
        "rate": body.rate,
        "weight": weight
    }

@app.post("/api/debit/{entry_id}/revert-rate-cut")
def revert_debit_rate_cut(entry_id: int, db: Session = Depends(get_db)):
    """Revert a Rate Cut payment back to a regular cash payment."""
    import re
    db_debit = db.query(models.DebitEntry).filter(models.DebitEntry.id == entry_id).first()
    if not db_debit:
        raise HTTPException(status_code=404, detail="Payment entry not found")
        
    clean_particulars = re.sub(r"\[RATE_CUT:[^\]]+\]\s*", "", db_debit.particulars or "")
    db_debit.particulars = clean_particulars.strip()
    db.commit()
    db.refresh(db_debit)
    
    if db_debit.daybook_id:
        crud.recalculate_and_cascade_daybook(db, db_debit.daybook_id)
        
    log_system_action(db, "REVERT_RATE_CUT", f"Reverted Rate Cut on Debit #{entry_id} (₹{db_debit.amount}) for supplier '{db_debit.name}' back to regular cash payment")
    return {"success": True, "particulars": db_debit.particulars}

# CREDIT
@app.post("/api/daybook/{daybook_id}/credit", response_model=schemas.CreditEntry)
def add_credit_entry(daybook_id: int, entry: schemas.CreditEntryCreate, db: Session = Depends(get_db)):
    daybook = crud.get_daybook(db, daybook_id)
    if not daybook:
        raise HTTPException(status_code=404, detail="DayBook not found")
    res = crud.create_credit_entry(db, entry, daybook_id)
    log_system_action(db, "CREDIT_CREATE", f"[{daybook.date}] Added Credit: {entry.name} | {entry.particulars} | ₹{entry.amount}")
    return res

@app.put("/api/credit/{entry_id}", response_model=schemas.CreditEntry)
def update_credit_entry(entry_id: int, entry_update: schemas.CreditEntryUpdate, db: Session = Depends(get_db)):
    res = crud.update_credit_entry(db, entry_id, entry_update)
    if not res:
        raise HTTPException(status_code=404, detail="Credit entry not found")
    log_system_action(db, "CREDIT_UPDATE", f"Updated Credit #{entry_id}: {res.name} | {res.particulars} | ₹{res.amount}")
    return res

@app.delete("/api/credit/{entry_id}")
def delete_credit_entry(entry_id: int, db: Session = Depends(get_db)):
    db_entry = db.query(models.CreditEntry).filter(models.CreditEntry.id == entry_id).first()
    if not db_entry:
        raise HTTPException(status_code=404, detail="Credit entry not found")
    details = f"Deleted Credit: {db_entry.name} | {db_entry.particulars} | ₹{db_entry.amount}"
    if not crud.delete_credit_entry(db, entry_id):
        raise HTTPException(status_code=404, detail="Credit entry not found")
    log_system_action(db, "CREDIT_DELETE", details)
@app.get("/api/udar/outstanding")
def get_outstanding_udhar(db: Session = Depends(get_db)):
    import re
    debit_entries = db.query(models.DebitEntry).all()
    credit_entries = db.query(models.CreditEntry).all()

    print("[DEBUG get_outstanding_udhar] DB URL:", db.bind.url, "| Debits count:", len(debit_entries), "| Credits count:", len(credit_entries))


    GENERIC_NAMES = {"udhar", "jama", "udhar/jama", "credit", "debit", "expenditure", "expense", "karcha"}

    def extract_person_name(name_str, part_str):
        name = (name_str or "").strip()
        part = (part_str or "").strip()

        if not name or name.lower() in GENERIC_NAMES:
            clean_part = re.sub(r"\[[^\]]+\]", "", part).strip()
            if clean_part:
                name = clean_part
        return name.strip()

    # Normalize key -> { display_name, amount }
    balances_map = {}

    for d in debit_entries:
        name_lower = (d.name or "").lower()
        part_lower = (d.particulars or "").lower()
        if "girvi no." in name_lower or "girvi pledge" in part_lower:
            continue
        c_name = extract_person_name(d.name, d.particulars)
        if not c_name:
            continue
        norm_key = c_name.lower()
        if norm_key not in balances_map:
            balances_map[norm_key] = {"name": c_name, "amount": 0.0}
        balances_map[norm_key]["amount"] += d.amount

    for c in credit_entries:
        name_lower = (c.name or "").lower()
        part_lower = (c.particulars or "").lower()
        if "chhudai no." in name_lower or "banda no." in name_lower or "girvi release" in part_lower or "girvi banda" in part_lower:
            continue
        c_name = extract_person_name(c.name, c.particulars)
        if not c_name:
            continue
        norm_key = c_name.lower()
        if norm_key not in balances_map:
            balances_map[norm_key] = {"name": c_name, "amount": 0.0}
        balances_map[norm_key]["amount"] -= c.amount

    res = []
    for item in balances_map.values():
        amt = round(item["amount"], 2)
        if abs(amt) > 0.01:
            res.append({"name": item["name"], "amount": amt})

    res.sort(key=lambda x: abs(x["amount"]), reverse=True)
    return res



# SOLD ITEM

@app.post("/api/daybook/{daybook_id}/sold-item", response_model=schemas.SoldItem)
def add_sold_item(daybook_id: int, item: schemas.SoldItemCreate, db: Session = Depends(get_db)):
    daybook = crud.get_daybook(db, daybook_id)
    if not daybook:
        raise HTTPException(status_code=404, detail="DayBook not found")
    res = crud.create_sold_item(db, item, daybook_id)
    log_system_action(db, "SALE_CREATE", f"[{daybook.date}] Sold Item: {item.item_name} | Qty: {item.quantity} | Wt: {item.weight}g | ₹{item.amount}")
    return res

@app.put("/api/sold-item/{item_id}", response_model=schemas.SoldItem)
def update_sold_item(item_id: int, item_update: schemas.SoldItemUpdate, db: Session = Depends(get_db)):
    res = crud.update_sold_item(db, item_id, item_update)
    if not res:
        raise HTTPException(status_code=404, detail="Sold item not found")
    log_system_action(db, "SALE_UPDATE", f"Updated Sale Item #{item_id}: {res.item_name} | Qty: {res.quantity} | Wt: {res.weight}g | ₹{res.amount}")
    return res

@app.delete("/api/sold-item/{item_id}")
def delete_sold_item(item_id: int, db: Session = Depends(get_db)):
    db_item = db.query(models.SoldItem).filter(models.SoldItem.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Sold item not found")
    details = f"Deleted Sale Item: {db_item.item_name} | Qty: {db_item.quantity} | Wt: {db_item.weight}g | ₹{db_item.amount}"
    if not crud.delete_sold_item(db, item_id):
        raise HTTPException(status_code=404, detail="Sold item not found")
    log_system_action(db, "SALE_DELETE", details)
    return {"message": "Sold item deleted"}

# PHONEPE / UPI
@app.post("/api/daybook/{daybook_id}/phonepe", response_model=schemas.PhonePeEntry)
def add_phonepe_entry(daybook_id: int, entry: schemas.PhonePeEntryCreate, db: Session = Depends(get_db)):
    daybook = crud.get_daybook(db, daybook_id)
    if not daybook:
        raise HTTPException(status_code=404, detail="DayBook not found")
    res = crud.create_phonepe_entry(db, entry, daybook_id)
    log_system_action(db, "PHONEPE_CREATE", f"[{daybook.date}] PhonePe (UPI) Entry: {entry.customer_name} | ₹{entry.amount}")
    return res

@app.delete("/api/phonepe/{entry_id}")
def delete_phonepe_entry(entry_id: int, db: Session = Depends(get_db)):
    db_entry = db.query(models.PhonePeEntry).filter(models.PhonePeEntry.id == entry_id).first()
    if not db_entry:
        raise HTTPException(status_code=404, detail="PhonePe entry not found")
    details = f"Deleted PhonePe Entry: {db_entry.customer_name} | ₹{db_entry.amount}"
    if not crud.delete_phonepe_entry(db, entry_id):
        raise HTTPException(status_code=404, detail="PhonePe entry not found")
    log_system_action(db, "PHONEPE_DELETE", details)
    return {"message": "PhonePe entry deleted"}

# OLD GOLD
@app.post("/api/daybook/{daybook_id}/old-gold", response_model=schemas.OldGoldEntry)
def add_old_gold_entry(daybook_id: int, entry: schemas.OldGoldEntryCreate, db: Session = Depends(get_db)):
    daybook = crud.get_daybook(db, daybook_id)
    if not daybook:
        raise HTTPException(status_code=404, detail="DayBook not found")
    res = crud.create_old_gold_entry(db, entry, daybook_id)
    log_system_action(db, "OLD_GOLD_CREATE", f"[{daybook.date}] Old Gold Received: {entry.customer_name} | Wt: {entry.weight}g | ₹{entry.amount}")
    return res

@app.delete("/api/old-gold/{entry_id}")
def delete_old_gold_entry(entry_id: int, db: Session = Depends(get_db)):
    db_entry = db.query(models.OldGoldEntry).filter(models.OldGoldEntry.id == entry_id).first()
    if not db_entry:
        raise HTTPException(status_code=404, detail="Old Gold entry not found")
    details = f"Deleted Old Gold: {db_entry.customer_name} | Wt: {db_entry.weight}g | ₹{db_entry.amount}"
    if not crud.delete_old_gold_entry(db, entry_id):
        raise HTTPException(status_code=404, detail="Old Gold entry not found")
    log_system_action(db, "OLD_GOLD_DELETE", details)
    return {"message": "Old Gold entry deleted"}

# OLD SILVER
@app.post("/api/daybook/{daybook_id}/old-silver", response_model=schemas.OldSilverEntry)
def add_old_silver_entry(daybook_id: int, entry: schemas.OldSilverEntryCreate, db: Session = Depends(get_db)):
    daybook = crud.get_daybook(db, daybook_id)
    if not daybook:
        raise HTTPException(status_code=404, detail="DayBook not found")
    res = crud.create_old_silver_entry(db, entry, daybook_id)
    log_system_action(db, "OLD_SILVER_CREATE", f"[{daybook.date}] Old Silver Received: {entry.customer_name} | Wt: {entry.weight}g | ₹{entry.amount}")
    return res

@app.delete("/api/old-silver/{entry_id}")
def delete_old_silver_entry(entry_id: int, db: Session = Depends(get_db)):
    db_entry = db.query(models.OldSilverEntry).filter(models.OldSilverEntry.id == entry_id).first()
    if not db_entry:
        raise HTTPException(status_code=404, detail="Old Silver entry not found")
    details = f"Deleted Old Silver: {db_entry.customer_name} | Wt: {db_entry.weight}g | ₹{db_entry.amount}"
    if not crud.delete_old_silver_entry(db, entry_id):
        raise HTTPException(status_code=404, detail="Old Silver entry not found")
    log_system_action(db, "OLD_SILVER_DELETE", details)
    return {"message": "Old Silver entry deleted"}


# PLEDGE
@app.post("/api/daybook/{daybook_id}/pledge", response_model=schemas.PledgeEntry)
def add_pledge_entry(daybook_id: int, entry: schemas.PledgeEntryCreate, db: Session = Depends(get_db)):
    daybook = crud.get_daybook(db, daybook_id)
    if not daybook:
        raise HTTPException(status_code=404, detail="DayBook not found")
    res = crud.create_pledge_entry(db, entry, daybook_id)
    log_system_action(db, "PLEDGE_CREATE", f"[{daybook.date}] Added Pledge: {entry.customer_name} | {entry.ornament} | Wt: {entry.weight}g | ₹{entry.amount}")
    return res

@app.delete("/api/pledge/{entry_id}")
def delete_pledge_entry(entry_id: int, db: Session = Depends(get_db)):
    db_entry = db.query(models.PledgeEntry).filter(models.PledgeEntry.id == entry_id).first()
    if not db_entry:
        raise HTTPException(status_code=404, detail="Pledge entry not found")
    details = f"Deleted Pledge: {db_entry.customer_name} | {db_entry.ornament} | Wt: {db_entry.weight}g | ₹{db_entry.amount}"
    if not crud.delete_pledge_entry(db, entry_id):
        raise HTTPException(status_code=404, detail="Pledge entry not found")
    log_system_action(db, "PLEDGE_DELETE", details)
    return {"message": "Pledge entry deleted"}

@app.get("/api/pledges", response_model=List[schemas.PledgeEntry])
def get_all_pledges(db: Session = Depends(get_db)):
    pledges = db.query(models.PledgeEntry).order_by(models.PledgeEntry.id.desc()).all()
    for p in pledges:
        p.date = p.daybook.date if p.daybook else None
    return pledges

@app.get("/api/sold-items")
def get_all_sold_items(db: Session = Depends(get_db)):
    items = db.query(models.SoldItem).order_by(models.SoldItem.id.desc()).all()
    return [
        {
            "id": item.id,
            "daybook_id": item.daybook_id,
            "item_name": item.item_name,
            "quantity": item.quantity,
            "weight": item.weight,
            "amount": item.amount,
            "date": item.daybook.date if item.daybook else ""
        } for item in items
    ]

@app.put("/api/pledge/{entry_id}", response_model=schemas.PledgeEntry)
def update_pledge_entry(entry_id: int, entry_update: schemas.PledgeEntryUpdate, db: Session = Depends(get_db)):
    db_entry = db.query(models.PledgeEntry).filter(models.PledgeEntry.id == entry_id).first()
    if not db_entry:
        raise HTTPException(status_code=404, detail="Pledge entry not found")

    updated_fields = entry_update.model_dump(exclude_unset=True)

    old_daybook_id = db_entry.daybook_id
    date_changed = False
    new_date = updated_fields.pop("date", None)
    if new_date:
        current_date = db_entry.daybook.date if db_entry.daybook else None
        if current_date and new_date != current_date:
            target_db = crud.get_daybook_by_date(db, new_date)
            if not target_db:
                prev = crud.get_previous_daybook(db, new_date)
                p_cash  = (prev.closing_cash or 0.0)  if prev else 0.0
                p_upi   = (prev.closing_upi or 0.0)   if prev else 0.0
                p_other = (prev.closing_other or 0.0) if prev else 0.0
                p_upi_details = (prev.closing_upi_details or "{}") if prev else "{}"
                target_db = crud.create_daybook(db, schemas.DayBookCreate(
                    date=new_date,
                    opening_cash=p_cash,
                    opening_upi=p_upi,
                    opening_other=p_other,
                    closing_cash=p_cash,
                    closing_upi=p_upi,
                    closing_other=p_other,
                    opening_upi_details=p_upi_details,
                    closing_upi_details=p_upi_details
                ))
            db_entry.daybook_id = target_db.id
            date_changed = True

            # If there's an upfront Banda credit entry, move it to the new daybook as well
            if db_entry.pledge_no:
                import re
                pledge_no = db_entry.pledge_no.strip()
                credit_pattern = rf"^banda no\.\s*{re.escape(pledge_no.lower())}$"
                credits = db.query(models.CreditEntry).filter(
                    models.CreditEntry.daybook_id == old_daybook_id,
                    models.CreditEntry.name.like(f"%{pledge_no}%")
                ).all()
                for c in credits:
                    if re.match(credit_pattern, c.name.lower().strip(), flags=re.IGNORECASE):
                        c.daybook_id = target_db.id

    for key, value in updated_fields.items():
        setattr(db_entry, key, value)

    if updated_fields.get("is_repledged") == 0:
        db_entry.is_repledged = 0
        db_entry.repledged_bank = None
        db_entry.repledged_amount = None
        db_entry.repledged_date = None
        db_entry.repledged_name = None
        db_entry.repledged_receipt_no = None
        db_entry.repledged_entries = None
        db_entry.repledged_interest_amount = None
        db_entry.repledged_interest_rate = None

    db.commit()
    db.refresh(db_entry)

    # Recalculate cash cascades when financial fields or date changed.
    FINANCIAL_FIELDS = {"amount", "status", "release_date", "is_existing", "is_repledged"}
    if date_changed or FINANCIAL_FIELDS.intersection(updated_fields.keys()):
        crud.recalculate_and_cascade_daybook(db, old_daybook_id)
        if date_changed:
            crud.recalculate_and_cascade_daybook(db, db_entry.daybook_id)

    db_entry.date = db_entry.daybook.date if db_entry.daybook else None
    log_system_action(db, "PLEDGE_UPDATE", f"Updated Pledge for {db_entry.customer_name} (ID: {entry_id}) to: {db_entry.ornament} | Wt: {db_entry.weight}g | ₹{db_entry.amount}")
    return db_entry

def number_to_words_inr(num: int) -> str:
    if num == 0:
        return "Zero"
    
    a = [
        "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
        "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"
    ]
    b = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    
    def convert_less_than_one_thousand(n: int) -> str:
        if n < 20:
            return a[n]
        digit = n % 10
        if n < 100:
            return b[n // 10] + (" " + a[digit] if digit else "")
        return a[n // 100] + " Hundred" + (" and " + convert_less_than_one_thousand(n % 100) if n % 100 != 0 else "")

    result = ""
    
    # Crores
    crores = num // 10000000
    remaining = num % 10000000
    if crores > 0:
        result += convert_less_than_one_thousand(crores) + " Crore "
        
    # Lakhs
    lakhs = remaining // 100000
    remaining = remaining % 100000
    if lakhs > 0:
        result += convert_less_than_one_thousand(lakhs) + " Lakh "
        
    # Thousands
    thousands = remaining // 1000
    remaining = remaining % 1000
    if thousands > 0:
        result += convert_less_than_one_thousand(thousands) + " Thousand "
        
    # Hundreds, tens, ones
    if remaining > 0:
        result += convert_less_than_one_thousand(remaining)
        
    return result.strip() + " only"

# PLEDGE PAYMENTS
@app.post("/api/pledge/{pledge_id}/payment", response_model=schemas.PledgePayment)
def add_pledge_payment(pledge_id: int, entry: schemas.PledgePaymentCreate, db: Session = Depends(get_db)):
    pledge = db.query(models.PledgeEntry).filter(models.PledgeEntry.id == pledge_id).first()
    if not pledge:
        raise HTTPException(status_code=404, detail="Pledge not found")
    
    # 1. Fetch or create daybook for date
    daybook = crud.get_daybook_by_date(db, entry.date)
    if not daybook:
        prev = crud.get_previous_daybook(db, entry.date)
        p_cash  = prev.closing_cash  if prev else 0.0
        p_upi   = prev.closing_upi   if prev else 0.0
        p_other = prev.closing_other if prev else 0.0
        p_upi_details = prev.closing_upi_details if (prev and prev.closing_upi_details) else "{}"
        
        new_daybook = schemas.DayBookCreate(
            date=entry.date,
            opening_cash=p_cash,
            opening_upi=p_upi,
            opening_other=p_other,
            closing_cash=p_cash,
            closing_upi=p_upi,
            closing_other=p_other,
            opening_upi_details=p_upi_details,
            closing_upi_details=p_upi_details
        )
        daybook = crud.create_daybook(db, new_daybook)
    
    # 2. Add payment record
    db_payment = models.PledgePayment(
        pledge_id=pledge_id,
        daybook_id=daybook.id,
        date=entry.date,
        payment_type=entry.payment_type,
        amount=entry.amount,
        payment_method=entry.payment_method
    )
    db.add(db_payment)
    db.commit()
    db.refresh(db_payment)
    
    # 3. Create entry to track cash flow and update pledge amounts
    if entry.payment_type == "TOP_UP":
        if entry.payment_method.startswith("UPI:"):
            acc = entry.payment_method.split(":")[1]
            prefix = f"[UPI:{acc}] "
        elif entry.payment_method == "UPI":
            prefix = "[UPI] "
        elif entry.payment_method == "OTHER":
            prefix = "[OTHER] "
        else:
            prefix = ""

        db_debit = models.DebitEntry(
            daybook_id=daybook.id,
            name=f"Girvi TopUp ({pledge.pledge_no})",
            particulars=f"{prefix}Girvi Principal Top-Up",
            amount=entry.amount,
            remarks=f"Principal Top-Up given to Pawner {pledge.customer_name} (Pledge {pledge.pledge_no})"
        )
        db.add(db_debit)
        
        # Update pledge amount
        pledge.amount += entry.amount
        pledge.rupees_in_words = number_to_words_inr(int(pledge.amount))
        db.commit()
    else:
        if entry.payment_method.startswith("UPI:"):
            acc = entry.payment_method.split(":")[1]
            prefix = f"[UPI:{acc}] "
        elif entry.payment_method == "UPI":
            prefix = "[UPI] "
        elif entry.payment_method == "OTHER":
            prefix = "[OTHER] "
        else:
            prefix = ""

        particulars = f"{prefix}Girvi interest payment" if entry.payment_type == "INTEREST" else f"{prefix}Girvi principal partial payment"

        db_credit = models.CreditEntry(
            daybook_id=daybook.id,
            name=f"Girvi Pay ({pledge.pledge_no})",
            particulars=particulars,
            amount=entry.amount,
            remarks=f"Payment for Pawner {pledge.customer_name} (Pledge {pledge.pledge_no})"
        )
        db.add(db_credit)
        db.commit()
    
    crud.recalculate_and_cascade_daybook(db, daybook.id)
    log_system_action(db, "PLEDGE_PAYMENT_CREATE", f"Recorded {entry.payment_type} Payment of ₹{entry.amount} for Pledge {pledge.pledge_no} via {entry.payment_method}")
    return db_payment

@app.delete("/api/pledge-payment/{payment_id}")
def delete_pledge_payment(payment_id: int, db: Session = Depends(get_db)):
    payment = db.query(models.PledgePayment).filter(models.PledgePayment.id == payment_id).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    pledge_no = payment.pledge.pledge_no if payment.pledge else "Unknown"
    
    if payment.payment_type == "TOP_UP":
        debit_name = f"Girvi TopUp ({pledge_no})"
        db_debit = db.query(models.DebitEntry).filter(
            models.DebitEntry.daybook_id == payment.daybook_id,
            models.DebitEntry.name == debit_name,
            models.DebitEntry.amount == payment.amount
        ).first()
        if db_debit:
            db.delete(db_debit)
        if payment.pledge:
            payment.pledge.amount -= payment.amount
            payment.pledge.rupees_in_words = number_to_words_inr(int(payment.pledge.amount))
    else:
        # Delete the matching CreditEntry in the daybook
        credit_name = f"Girvi Pay ({pledge_no})"
        db_credit = db.query(models.CreditEntry).filter(
            models.CreditEntry.daybook_id == payment.daybook_id,
            models.CreditEntry.name == credit_name,
            models.CreditEntry.amount == payment.amount
        ).first()
        if db_credit:
            db.delete(db_credit)
    
    db_daybook_id = payment.daybook_id
    db.delete(payment)
    db.commit()
    
    crud.recalculate_and_cascade_daybook(db, db_daybook_id)
    log_system_action(db, "PLEDGE_PAYMENT_DELETE", f"Deleted payment (ID: {payment_id}) of ₹{payment.amount} for Pledge {pledge_no}")
    return {"message": "Payment deleted successfully"}

# RELEASE
@app.post("/api/daybook/{daybook_id}/release", response_model=schemas.ReleaseEntry)
def add_release_entry(daybook_id: int, entry: schemas.ReleaseEntryCreate, db: Session = Depends(get_db)):
    daybook = crud.get_daybook(db, daybook_id)
    if not daybook:
        raise HTTPException(status_code=404, detail="DayBook not found")
    res = crud.create_release_entry(db, entry, daybook_id)
    log_system_action(db, "PLEDGE_RELEASE", f"[{daybook.date}] Released Pledge: {entry.customer_name} | Principal: ₹{entry.principal_amount} | Interest: ₹{entry.interest_received}")
    return res

@app.delete("/api/release/{entry_id}")
def delete_release_entry(entry_id: int, db: Session = Depends(get_db)):
    db_entry = db.query(models.ReleaseEntry).filter(models.ReleaseEntry.id == entry_id).first()
    if not db_entry:
        raise HTTPException(status_code=404, detail="Release entry not found")
    details = f"Deleted Pledge Release: {db_entry.customer_name} | Principal: ₹{db_entry.principal_amount} | Interest: ₹{db_entry.interest_received}"
    if not crud.delete_release_entry(db, entry_id):
        raise HTTPException(status_code=404, detail="Release entry not found")
    log_system_action(db, "RELEASE_DELETE", details)
    return {"message": "Release entry deleted"}

@app.post("/api/pledge/{pledge_id}/revert-release", response_model=schemas.PledgeEntry)
def revert_pledge_release(pledge_id: int, db: Session = Depends(get_db)):
    db_entry = db.query(models.PledgeEntry).filter(models.PledgeEntry.id == pledge_id).first()
    if not db_entry:
        raise HTTPException(status_code=404, detail="Pledge entry not found")

    if db_entry.status != "RELEASED":
        raise HTTPException(status_code=400, detail="Pledge is not in RELEASED status")

    affected_daybook_ids = set()
    if db_entry.daybook_id:
        affected_daybook_ids.add(db_entry.daybook_id)

    pledge_no = (db_entry.pledge_no or "").strip()

    # Revert pledge status to ACTIVE and release_date to None
    db_entry.status = "ACTIVE"
    db_entry.release_date = None

    # Delete corresponding ReleaseEntry records in daybook
    if pledge_no:
        import re
        releases = db.query(models.ReleaseEntry).all()
        for r in releases:
            cleaned_name = re.sub(r"^\[[^\]]+\]\s*", "", r.customer_name, flags=re.IGNORECASE).strip()
            if cleaned_name == pledge_no or r.customer_name == pledge_no or r.customer_name.endswith(f" {pledge_no}"):
                affected_daybook_ids.add(r.daybook_id)
                db.delete(r)

    db.commit()

    for db_id in affected_daybook_ids:
        crud.recalculate_and_cascade_daybook(db, db_id)

    db.refresh(db_entry)
    db_entry.date = db_entry.daybook.date if db_entry.daybook else None
    log_system_action(db, "PLEDGE_REVERT_RELEASE", f"Reverted release for Pledge #{db_entry.pledge_no} ({db_entry.customer_name}) back to ACTIVE")
    return db_entry



# Live Gold & Silver rates Bangalore scraper
import time
import urllib.request
import re

LIVE_RATES_CACHE = {
    "gold_24k": "14,433",
    "gold_22k": "13,230",
    "gold_18k": "10,825",
    "silver": "2,45,000",
    "last_updated": 0
}

@app.get("/api/live-rates")
def get_live_rates():
    global LIVE_RATES_CACHE
    now = time.time()
    # Cache for 1 hour (3600 seconds)
    if now - LIVE_RATES_CACHE["last_updated"] > 3600:
        url = "https://www.goodreturns.in/gold-rates/bangalore.html"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        req = urllib.request.Request(url, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=5) as response:
                html = response.read().decode('utf-8')
                
                match_24 = re.search(r'&#x20b9;([\d,]+)</strong>\s*per gram for 24 karat gold', html, re.IGNORECASE)
                match_22 = re.search(r'&#x20b9;([\d,]+)</strong>\s*per gram for 22 karat gold', html, re.IGNORECASE)
                match_18 = re.search(r'&#x20b9;([\d,]+)</strong>\s*per gram for 18 karat gold', html, re.IGNORECASE)
                match_silver = re.search(r'<span class="label">Silver</span>\s*<span class="value">[^<]*?([\d,]+)\s*/\s*kg', html, re.IGNORECASE)
                
                if match_24:
                    LIVE_RATES_CACHE["gold_24k"] = match_24.group(1)
                if match_22:
                    LIVE_RATES_CACHE["gold_22k"] = match_22.group(1)
                if match_18:
                    LIVE_RATES_CACHE["gold_18k"] = match_18.group(1)
                if match_silver:
                    LIVE_RATES_CACHE["silver"] = match_silver.group(1)
                    
                LIVE_RATES_CACHE["last_updated"] = now
        except Exception as e:
            print("Failed to fetch live rates, using cached/default rates:", e)
            
    return {
        "gold_24k": LIVE_RATES_CACHE["gold_24k"],
        "gold_22k": LIVE_RATES_CACHE["gold_22k"],
        "gold_18k": LIVE_RATES_CACHE["gold_18k"],
        "silver": LIVE_RATES_CACHE["silver"]
    }

# Barcode cache & reload mechanism
import os
import openpyxl

BARCODE_CACHE = {}
LAST_LOADED_GOLD_MTIME = 0
LAST_LOADED_SILVER_MTIME = 0

def load_barcode_cache():
    global BARCODE_CACHE, LAST_LOADED_GOLD_MTIME, LAST_LOADED_SILVER_MTIME
    
    # Paths to Excel files (workspace root, since files are in "DAY BOOK/DAY BOOK")
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    gold_path = os.path.join(base_dir, "GOLD.xlsx")
    silver_path = os.path.join(base_dir, "SILVER.xlsx")
    
    # Check modification times
    gold_mtime = os.path.getmtime(gold_path) if os.path.exists(gold_path) else 0
    silver_mtime = os.path.getmtime(silver_path) if os.path.exists(silver_path) else 0
    
    if gold_mtime != LAST_LOADED_GOLD_MTIME or silver_mtime != LAST_LOADED_SILVER_MTIME:
        print("[Barcode Cache] Reloading barcode database...")
        new_cache = {}
        
        # Load GOLD.xlsx
        if os.path.exists(gold_path):
            try:
                wb = openpyxl.load_workbook(gold_path, read_only=True)
                sheet = wb.active
                # Headers are on row index 4 (1-based row index 5)
                # Data starts from row index 5 (1-based row index 6)
                rows = list(sheet.iter_rows(values_only=True))
                if len(rows) > 5:
                    headers = [str(h).strip() if h is not None else "" for h in rows[4]]
                    try:
                        barcode_idx = headers.index("BarcodeNo")
                        name_idx = headers.index("OrnamentName")
                        purity_idx = headers.index("Purity")
                        qty_idx = headers.index("Qty")
                        wt_idx = headers.index("Wt")
                        huid_idx = headers.index("HUIDNo") if "HUIDNo" in headers else -1
                        
                        for row in rows[5:]:
                            if len(row) > max(barcode_idx, name_idx, purity_idx, qty_idx, wt_idx):
                                b_val = row[barcode_idx]
                                if b_val:
                                    b_str = str(b_val).strip()
                                    new_cache[b_str] = {
                                        "metal": "GOLD",
                                        "itemName": str(row[name_idx]).strip() if row[name_idx] is not None else "",
                                        "purity": str(row[purity_idx]).strip() if row[purity_idx] is not None else "",
                                        "qty": int(row[qty_idx]) if row[qty_idx] is not None else 1,
                                        "weight": float(row[wt_idx]) if row[wt_idx] is not None else 0.0,
                                        "huid": str(row[huid_idx]).strip() if (huid_idx != -1 and row[huid_idx] is not None) else "",
                                    }
                    except ValueError as ve:
                        print(f"[Barcode Cache] Missing column headers in GOLD.xlsx: {ve}")
            except Exception as e:
                print(f"[Barcode Cache] Error loading GOLD.xlsx: {e}")
        
        # Load SILVER.xlsx
        if os.path.exists(silver_path):
            try:
                wb = openpyxl.load_workbook(silver_path, read_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
                if len(rows) > 5:
                    headers = [str(h).strip() if h is not None else "" for h in rows[4]]
                    try:
                        barcode_idx = headers.index("BarcodeNo")
                        name_idx = headers.index("OrnamentName")
                        purity_idx = headers.index("Purity")
                        qty_idx = headers.index("Qty")
                        wt_idx = headers.index("Wt")
                        huid_idx = headers.index("HUIDNo") if "HUIDNo" in headers else -1
                        
                        for row in rows[5:]:
                            if len(row) > max(barcode_idx, name_idx, purity_idx, qty_idx, wt_idx):
                                b_val = row[barcode_idx]
                                if b_val:
                                    b_str = str(b_val).strip()
                                    new_cache[b_str] = {
                                        "metal": "SILVER",
                                        "itemName": str(row[name_idx]).strip() if row[name_idx] is not None else "",
                                        "purity": str(row[purity_idx]).strip() if row[purity_idx] is not None else "",
                                        "qty": int(row[qty_idx]) if row[qty_idx] is not None else 1,
                                        "weight": float(row[wt_idx]) if row[wt_idx] is not None else 0.0,
                                        "huid": str(row[huid_idx]).strip() if (huid_idx != -1 and row[huid_idx] is not None) else "",
                                    }
                    except ValueError as ve:
                        print(f"[Barcode Cache] Missing column headers in SILVER.xlsx: {ve}")
            except Exception as e:
                print(f"[Barcode Cache] Error loading SILVER.xlsx: {e}")
        
        BARCODE_CACHE = new_cache
        LAST_LOADED_GOLD_MTIME = gold_mtime
        LAST_LOADED_SILVER_MTIME = silver_mtime
        print(f"[Barcode Cache] Loaded {len(BARCODE_CACHE)} items.")

@app.get("/api/barcode/{barcode_no}")
def get_barcode_item(barcode_no: str, db: Session = Depends(get_db)):
    try:
        load_barcode_cache()
    except Exception as e:
        print(f"[Barcode Cache] Reload failed: {e}")
    
    bc_clean = barcode_no.strip().upper()
    
    # 1. Determine if already sold and what date
    is_sold = False
    sold_date = None
    
    # Query database SoldItem records containing this barcode
    sold_item = db.query(models.SoldItem).filter(
        models.SoldItem.item_name.like(f"%[BARCODE:{bc_clean}]%")
    ).first()
    
    if sold_item:
        is_sold = True
        sold_date = sold_item.daybook.date if sold_item.daybook else "Unknown Date"
    else:
        # Check purchase items is_sold column
        db_item_check = db.query(models.PurchaseItem).filter(
            models.PurchaseItem.barcode_no == bc_clean
        ).first()
        if db_item_check and getattr(db_item_check, 'is_sold', 0) == 1:
            is_sold = True
            sold_item_check = db.query(models.SoldItem).filter(
                models.SoldItem.item_name.like(f"%[BARCODE:{bc_clean}]%")
            ).first()
            sold_date = sold_item_check.daybook.date if (sold_item_check and sold_item_check.daybook) else "Stock Database"
        else:
            # Check excel sold overrides
            excel_sold = db.query(models.SoldExcelBarcode).filter(
                models.SoldExcelBarcode.barcode_no == bc_clean
            ).first()
            if excel_sold:
                is_sold = True
                sold_item_check = db.query(models.SoldItem).filter(
                    models.SoldItem.item_name.like(f"%[BARCODE:{bc_clean}]%")
                ).first()
                sold_date = sold_item_check.daybook.date if (sold_item_check and sold_item_check.daybook) else "Excel Override"

    # 2. Check xlsx cache first
    item = BARCODE_CACHE.get(bc_clean)
    if item:
        return {
            "found": True,
            "is_sold": is_sold,
            "sold_date": sold_date,
            "item": item
        }

    # 3. Fallback: check purchase_items table in DB
    db_item = db.query(models.PurchaseItem).filter(
        models.PurchaseItem.barcode_no == bc_clean
    ).first()
    if db_item:
        return {
            "found": True,
            "is_sold": is_sold,
            "sold_date": sold_date,
            "item": {
                "metal": db_item.metal or "GOLD",
                "itemName": db_item.ornament_name,
                "purity": db_item.purity,
                "qty": db_item.qty,
                "weight": db_item.weight,
                "huid": db_item.huid_no or "",
            }
        }

    return {"found": False}


# ── Purchase Bill routes ────────────────────────────────────────────────────

from pydantic import BaseModel as PydanticBase
from typing import List as TypingList

class PurchaseItemIn(PydanticBase):
    ornament_name: str
    huid_no: str = ""
    purity: str
    qty: int = 1
    weight: float
    net_weight: float = 0.0
    rate: float = 0.0
    making: str = ""
    amount: float = 0.0
    remark: str = ""
    barcode_no: str

class PurchaseBillIn(PydanticBase):
    bill_no: str
    bill_date: str
    supplier_name: str
    supplier_gst: str = ""
    metal: str = "GOLD"
    invoice_total: float = 0.0
    remarks: str = ""
    total_weight: Optional[float] = None
    purity: Optional[str] = None
    is_rate_cut: Optional[int] = 1  # 1 = Yes, 0 = No
    rate: Optional[float] = None
    amount: Optional[float] = None
    pure_weight: Optional[float] = None
    making: Optional[float] = None
    total_percent: Optional[float] = None
    items: TypingList[PurchaseItemIn]

@app.get("/api/purchase/next-barcode")
def get_next_barcode(db: Session = Depends(get_db)):
    """Return the next available P-series barcode number (P00001, P00002, …)."""
    max_num = 0
    db_barcodes = db.query(models.PurchaseItem.barcode_no).all()
    for (bc,) in db_barcodes:
        if bc and bc.upper().startswith("P"):
            try:
                n = int(bc[1:])
                if n > max_num:
                    max_num = n
            except ValueError:
                pass

    return {"next_barcode_number": max_num + 1}

@app.post("/api/purchase/bill")
def create_purchase_bill(bill: PurchaseBillIn, db: Session = Depends(get_db)):
    """Save a new purchase bill and all its items."""
    import datetime

    # Check for duplicate bill_no
    existing = db.query(models.PurchaseBill).filter(
        models.PurchaseBill.bill_no == bill.bill_no
    ).first()
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Purchase bill '{bill.bill_no}' already exists."
        )

    # Check for duplicate barcodes
    for it in bill.items:
        dup_xlsx = BARCODE_CACHE.get(it.barcode_no.strip())
        if dup_xlsx:
            raise HTTPException(
                status_code=400,
                detail=f"Barcode {it.barcode_no} already exists in stock (xlsx)."
            )
        dup_db = db.query(models.PurchaseItem).filter(
            models.PurchaseItem.barcode_no == it.barcode_no.strip()
        ).first()
        if dup_db:
            raise HTTPException(
                status_code=400,
                detail=f"Barcode {it.barcode_no} already exists in database."
            )

    now_str = datetime.datetime.now().isoformat()
    db_bill = models.PurchaseBill(
        bill_no=bill.bill_no,
        bill_date=bill.bill_date,
        supplier_name=bill.supplier_name,
        supplier_gst=bill.supplier_gst,
        metal=bill.metal,
        invoice_total=bill.invoice_total,
        remarks=bill.remarks,
        created_at=now_str,
        total_weight=bill.total_weight,
        purity=bill.purity,
        is_rate_cut=bill.is_rate_cut if bill.is_rate_cut is not None else 1,
        rate=bill.rate,
        amount=bill.amount,
        pure_weight=bill.pure_weight,
        making=bill.making,
        total_percent=bill.total_percent,
    )
    db.add(db_bill)
    db.flush()  # get db_bill.id

    saved_items = []
    for it in bill.items:
        db_item = models.PurchaseItem(
            bill_id=db_bill.id,
            barcode_no=it.barcode_no.strip(),
            ornament_name=it.ornament_name,
            huid_no=it.huid_no,
            purity=it.purity,
            qty=it.qty,
            weight=it.weight,
            net_weight=it.net_weight if it.net_weight else it.weight,
            rate=it.rate,
            making=it.making,
            amount=it.amount,
            remark=it.remark,
            bill_no=bill.bill_no,
            bill_date=bill.bill_date,
            metal=bill.metal,
        )
        db.add(db_item)
        saved_items.append({
            "barcode_no": it.barcode_no,
            "ornament_name": it.ornament_name,
            "purity": it.purity,
            "weight": it.weight,
            "qty": it.qty,
            "amount": it.amount,
        })

    db.commit()
    log_system_action(db, "PURCHASE_BILL_CREATE", f"Saved Purchase Bill: No: {db_bill.bill_no} | Supplier: {db_bill.supplier_name} | Metal: {db_bill.metal} | Total: ₹{db_bill.invoice_total}")

    return {
        "id": db_bill.id,
        "bill_no": db_bill.bill_no,
        "bill_date": db_bill.bill_date,
        "supplier_name": db_bill.supplier_name,
        "metal": db_bill.metal,
        "invoice_total": db_bill.invoice_total,
        "items": saved_items,
        "created_at": now_str,
        "total_weight": db_bill.total_weight,
        "purity": db_bill.purity,
        "is_rate_cut": db_bill.is_rate_cut,
        "rate": db_bill.rate,
        "amount": db_bill.amount,
        "pure_weight": db_bill.pure_weight,
        "making": db_bill.making,
        "total_percent": db_bill.total_percent,
    }


@app.get("/api/purchase/suppliers")
def get_unique_suppliers(db: Session = Depends(get_db)):
    """Fetch distinct supplier names and GST numbers from both existing purchase bills and registered parties."""
    # 1. From registered purchase parties
    party_results = db.query(
        models.PurchaseParty.name,
        models.PurchaseParty.gstin
    ).all()
    
    # 2. From existing purchase bills
    bill_results = db.query(
        models.PurchaseBill.supplier_name,
        models.PurchaseBill.supplier_gst
    ).distinct().all()
    
    unique_suppliers = []
    seen_names = set()
    
    # Add registered suppliers first
    for name, gst in party_results:
        if name and name.strip():
            norm_name = name.strip().lower()
            if norm_name not in seen_names:
                seen_names.add(norm_name)
                unique_suppliers.append({
                    "supplier_name": name.strip(),
                    "supplier_gst": gst.strip() if gst else ""
                })
                
    # Add virtual/bill suppliers
    for name, gst in bill_results:
        if name and name.strip():
            norm_name = name.strip().lower()
            if norm_name not in seen_names:
                seen_names.add(norm_name)
                unique_suppliers.append({
                    "supplier_name": name.strip(),
                    "supplier_gst": gst.strip() if gst else ""
                })
                
    unique_suppliers.sort(key=lambda x: x["supplier_name"].lower())
    return unique_suppliers


@app.get("/api/purchase/bills")
def list_purchase_bills(db: Session = Depends(get_db)):
    bills = db.query(models.PurchaseBill).order_by(models.PurchaseBill.id.desc()).all()
    return [
        {
            "id": b.id,
            "bill_no": b.bill_no,
            "bill_date": b.bill_date,
            "supplier_name": b.supplier_name,
            "metal": b.metal,
            "invoice_total": b.invoice_total,
            "item_count": len(b.items),
            "created_at": b.created_at,
            "total_weight": b.total_weight,
            "purity": b.purity,
            "is_rate_cut": b.is_rate_cut,
            "rate": b.rate,
            "amount": b.amount,
            "pure_weight": b.pure_weight,
            "making": b.making,
            "total_percent": b.total_percent,
        }
        for b in bills
    ]

@app.get("/api/purchase/bill/{bill_id}")
def get_purchase_bill(bill_id: int, db: Session = Depends(get_db)):
    b = db.query(models.PurchaseBill).filter(models.PurchaseBill.id == bill_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Purchase bill not found")
    return {
        "id": b.id,
        "bill_no": b.bill_no,
        "bill_date": b.bill_date,
        "supplier_name": b.supplier_name,
        "supplier_gst": b.supplier_gst,
        "metal": b.metal,
        "invoice_total": b.invoice_total,
        "remarks": b.remarks,
        "created_at": b.created_at,
        "total_weight": b.total_weight,
        "purity": b.purity,
        "is_rate_cut": b.is_rate_cut,
        "rate": b.rate,
        "amount": b.amount,
        "pure_weight": b.pure_weight,
        "making": b.making,
        "total_percent": b.total_percent,
        "items": [
            {
                "id": it.id,
                "barcode_no": it.barcode_no,
                "ornament_name": it.ornament_name,
                "huid_no": it.huid_no,
                "purity": it.purity,
                "qty": it.qty,
                "weight": it.weight,
                "net_weight": it.net_weight,
                "rate": it.rate,
                "making": it.making,
                "amount": it.amount,
                "remark": it.remark,
            }
            for it in b.items
        ],
    }

@app.delete("/api/purchase/bill/{bill_id}")
def delete_purchase_bill(bill_id: int, db: Session = Depends(get_db)):
    b = db.query(models.PurchaseBill).filter(models.PurchaseBill.id == bill_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Purchase bill not found")
    details = f"Deleted Purchase Bill: No: {b.bill_no} | Supplier: {b.supplier_name} | Metal: {b.metal} | Total: ₹{b.invoice_total}"
    db.delete(b)
    db.commit()
    log_system_action(db, "PURCHASE_BILL_DELETE", details)
    return {"message": "Purchase bill deleted"}


# ── Inventory & Stock Register routes ──────────────────────────────────────────

@app.get("/api/inventory")
def get_inventory(
    metal: Optional[str] = None,
    purity: Optional[str] = None,
    status: Optional[str] = None,  # "available", "sold", "all"
    search: Optional[str] = None,
    source: Optional[str] = None,  # "EXCEL", "DB", "ALL"
    db: Session = Depends(get_db)
):
    try:
        load_barcode_cache()
    except Exception as e:
        print(f"[Inventory] Cache load failed: {e}")

    # Build dictionary of all sold barcodes -> sold_date
    sold_barcodes = {}
    
    # 1. From sold_items (real daybook sales)
    import re
    sold_items = db.query(models.SoldItem).all()
    for si in sold_items:
        if si.item_name:
            match = re.search(r"\[BARCODE:([A-Za-z0-9]+)\]", si.item_name)
            if match:
                bc_key = match.group(1).strip().upper()
                s_date = si.daybook.date if si.daybook else ""
                sold_barcodes[bc_key] = s_date
                
    # 2. From sold_excel_barcodes (manual overrides for excel items)
    manual_sold = db.query(models.SoldExcelBarcode).all()
    for seb in manual_sold:
        if seb.barcode_no:
            bc_key = seb.barcode_no.strip().upper()
            if bc_key not in sold_barcodes:
                sold_barcodes[bc_key] = seb.sold_date or ""

    items_list = []

    # Process Excel Cache items
    if source in (None, "ALL", "EXCEL"):
        for bc, details in BARCODE_CACHE.items():
            bc_upper = bc.strip().upper()
            is_sold = 1 if bc_upper in sold_barcodes else 0
            sold_date = sold_barcodes.get(bc_upper, "") if is_sold == 1 else ""
            
            # Apply filters
            if metal and details.get("metal") != metal:
                continue
            if purity and details.get("purity") != purity:
                continue
            if status == "available" and is_sold == 1:
                continue
            if status == "sold" and is_sold == 0:
                continue
            
            # Search query (matches barcode or item name)
            if search:
                s_lower = search.lower()
                name_match = s_lower in details.get("itemName", "").lower()
                bc_match = s_lower in bc.lower()
                if not (name_match or bc_match):
                    continue

            items_list.append({
                "barcode_no": bc,
                "ornament_name": details.get("itemName", ""),
                "metal": details.get("metal", "GOLD"),
                "purity": details.get("purity", ""),
                "qty": details.get("qty", 1),
                "weight": details.get("weight", 0.0),
                "net_weight": details.get("weight", 0.0),
                "huid_no": details.get("huid", ""),
                "is_sold": is_sold,
                "sold_date": sold_date,
                "source": "EXCEL",
                "bill_no": "Pre-Existing Stock",
                "bill_date": "",
            })

    # Process DB items
    if source in (None, "ALL", "DB"):
        db_items = db.query(models.PurchaseItem).all()
        for pi in db_items:
            bc_upper = pi.barcode_no.strip().upper()
            is_sold = 1 if (pi.is_sold == 1 or bc_upper in sold_barcodes) else 0
            sold_date = (sold_barcodes.get(bc_upper) or pi.sold_date or "") if is_sold == 1 else ""

            # Apply filters
            if metal and pi.metal != metal:
                continue
            if purity and pi.purity != purity:
                continue
            if status == "available" and is_sold == 1:
                continue
            if status == "sold" and is_sold == 0:
                continue

            if search:
                s_lower = search.lower()
                name_match = s_lower in pi.ornament_name.lower()
                bc_match = s_lower in pi.barcode_no.lower()
                if not (name_match or bc_match):
                    continue

            items_list.append({
                "barcode_no": pi.barcode_no,
                "ornament_name": pi.ornament_name,
                "metal": pi.metal or "GOLD",
                "purity": pi.purity,
                "qty": pi.qty,
                "weight": pi.weight,
                "net_weight": pi.net_weight or pi.weight,
                "huid_no": pi.huid_no or "",
                "is_sold": is_sold,
                "sold_date": sold_date,
                "source": "DB",
                "bill_no": pi.bill_no or "—",
                "bill_date": pi.bill_date or "",
            })

    return items_list

class ToggleSoldRequest(BaseModel):
    sold_date: Optional[str] = None

@app.post("/api/inventory/{barcode_no}/toggle-sold")
def toggle_sold_status(barcode_no: str, payload: Optional[ToggleSoldRequest] = None, db: Session = Depends(get_db)):
    barcode_no = barcode_no.strip()
    bc_upper = barcode_no.upper()
    
    from datetime import datetime
    date_to_use = (payload.sold_date.strip() if (payload and payload.sold_date) else "") or datetime.now().strftime("%Y-%m-%d")

    # 1. Toggle for DB item if it exists
    db_item = db.query(models.PurchaseItem).filter(models.PurchaseItem.barcode_no == barcode_no).first()
    if db_item:
        if db_item.is_sold == 1:
            db_item.is_sold = 0
            db_item.sold_date = None
        else:
            db_item.is_sold = 1
            db_item.sold_date = date_to_use
        db.commit()
        return {"success": True, "is_sold": db_item.is_sold, "sold_date": db_item.sold_date or ""}

    # 2. Toggle for Excel item (check if it exists in cache)
    try:
        load_barcode_cache()
    except Exception:
        pass

    if barcode_no in BARCODE_CACHE:
        excel_sold = db.query(models.SoldExcelBarcode).filter(models.SoldExcelBarcode.barcode_no == barcode_no).first()
        if excel_sold:
            db.delete(excel_sold)
            db.commit()
            return {"success": True, "is_sold": 0, "sold_date": ""}
        else:
            new_sold = models.SoldExcelBarcode(barcode_no=barcode_no, sold_date=date_to_use)
            db.add(new_sold)
            db.commit()
            return {"success": True, "is_sold": 1, "sold_date": date_to_use}

    raise HTTPException(status_code=404, detail="Barcode not found in stock lists")


# --- Purchase Party / Supplier API Endpoints ---

@app.get("/api/purchase/parties")
def get_purchase_parties(db: Session = Depends(get_db)):
    from sqlalchemy import func
    import re
    
    def parse_rate_cut_payment(particulars: str):
        match = re.search(r"\[RATE_CUT:(GOLD|SILVER)\|([\d.]+)\|([\d.]+)\]", particulars)
        if match:
            metal = match.group(1)
            rate = float(match.group(2))
            weight = float(match.group(3))
            return metal, rate, weight
        return None

    # 1. Fetch all registered parties
    parties = db.query(models.PurchaseParty).order_by(models.PurchaseParty.name.asc()).all()
    
    # 2. Fetch distinct supplier names from bills to include virtual parties
    results = db.query(
        models.PurchaseBill.supplier_name,
        models.PurchaseBill.supplier_gst
    ).distinct().all()
    
    party_map = {p.name.strip().lower(): p for p in parties}
    
    merged_parties = []
    # Add registered parties
    for p in parties:
        merged_parties.append({
            "id": p.id,
            "name": p.name.strip(),
            "phone": p.phone or "",
            "gstin": p.gstin or "",
            "opening_balance_cash": p.opening_balance_cash or 0.0,
            "opening_balance_gold": p.opening_balance_gold or 0.0,
            "opening_balance_silver": p.opening_balance_silver or 0.0,
            "address": p.address or "",
            "created_at": p.created_at or ""
        })
        
    # Add virtual parties from bills that aren't registered yet
    for name, gst in results:
        if name and name.strip():
            norm_name = name.strip().lower()
            if norm_name not in party_map:
                merged_parties.append({
                    "id": 0,  # 0 indicates unregistered/virtual party
                    "name": name.strip(),
                    "phone": "",
                    "gstin": gst.strip() if gst else "",
                    "opening_balance_cash": 0.0,
                    "opening_balance_gold": 0.0,
                    "opening_balance_silver": 0.0,
                    "address": "",
                    "created_at": ""
                })
                party_map[norm_name] = None
                
    # Recalculate balances dynamically for each party
    for p_data in merged_parties:
        name_lower = p_data["name"].lower()
        
        # Fetch bills
        bills = db.query(models.PurchaseBill).filter(func.lower(models.PurchaseBill.supplier_name) == name_lower).all()
        total_bill_amount = sum(b.invoice_total or 0.0 for b in bills)
        
        # Fetch debits (payments)
        debits = db.query(models.DebitEntry).filter(func.lower(models.DebitEntry.name) == name_lower).all()
        
        total_regular_paid = 0.0
        gold_paid_wt = 0.0
        silver_paid_wt = 0.0
        
        for d in debits:
            rate_cut = parse_rate_cut_payment(d.particulars)
            if rate_cut:
                metal, rate, weight = rate_cut
                if metal == "GOLD":
                    gold_paid_wt += weight
                elif metal == "SILVER":
                    silver_paid_wt += weight
            else:
                total_regular_paid += d.amount
                
        p_data["outstanding_cash"] = p_data["opening_balance_cash"] + total_bill_amount - total_regular_paid
        
        # Gold Outstanding Balance (grams)
        # We owe: opening + pure weight from GOLD bills - gold weight paid via rate-cut payments
        gold_bills = db.query(models.PurchaseBill).filter(
            func.lower(models.PurchaseBill.supplier_name) == name_lower,
            models.PurchaseBill.metal == "GOLD",
            models.PurchaseBill.is_rate_cut == 0
        ).all()
        total_gold_wt = sum(b.pure_weight or 0.0 for b in gold_bills)
        p_data["outstanding_gold"] = p_data["opening_balance_gold"] + total_gold_wt - gold_paid_wt
        
        # Silver Outstanding Balance (grams)
        # We owe: opening + pure weight from SILVER bills - silver weight paid via rate-cut payments
        silver_bills = db.query(models.PurchaseBill).filter(
            func.lower(models.PurchaseBill.supplier_name) == name_lower,
            models.PurchaseBill.metal == "SILVER",
            models.PurchaseBill.is_rate_cut == 0
        ).all()
        total_silver_wt = sum(b.pure_weight or 0.0 for b in silver_bills)
        p_data["outstanding_silver"] = p_data["opening_balance_silver"] + total_silver_wt - silver_paid_wt
        
    merged_parties.sort(key=lambda x: x["name"].lower())
    return merged_parties


@app.post("/api/purchase/parties", response_model=schemas.PurchasePartyResponse)
def create_purchase_party(party: schemas.PurchasePartyCreate, db: Session = Depends(get_db)):
    from sqlalchemy import func
    normalized_name = party.name.strip().lower()
    existing = db.query(models.PurchaseParty).filter(func.lower(models.PurchaseParty.name) == normalized_name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Supplier with this name already exists")
        
    import datetime
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    db_party = models.PurchaseParty(
        name=party.name.strip(),
        phone=party.phone.strip() if party.phone else None,
        gstin=party.gstin.strip().upper() if party.gstin else None,
        opening_balance_cash=party.opening_balance_cash,
        opening_balance_gold=party.opening_balance_gold,
        opening_balance_silver=party.opening_balance_silver,
        address=party.address.strip() if party.address else None,
        created_at=now_str
    )
    db.add(db_party)
    db.commit()
    db.refresh(db_party)
    log_system_action(db, "PARTY_CREATE", f"Registered Supplier: {db_party.name} | Phone: {db_party.phone} | Balances: cash ₹{db_party.opening_balance_cash}, gold {db_party.opening_balance_gold}g, silver {db_party.opening_balance_silver}g")
    return db_party


@app.get("/api/purchase/parties/{party_name}/transactions")
def get_party_transactions(party_name: str, db: Session = Depends(get_db)):
    from sqlalchemy import func
    import re
    
    def parse_rate_cut_payment(particulars: str):
        match = re.search(r"\[RATE_CUT:(GOLD|SILVER)\|([\d.]+)\|([\d.]+)\]", particulars)
        if match:
            metal = match.group(1)
            rate = float(match.group(2))
            weight = float(match.group(3))
            return metal, rate, weight
        return None

    normalized_name = party_name.strip().lower()
    
    # Fetch bills for the supplier
    bills = db.query(models.PurchaseBill).filter(func.lower(models.PurchaseBill.supplier_name) == normalized_name).all()
    
    # Fetch payments (Debit entries) for the supplier
    debits = db.query(models.DebitEntry, models.DayBook.date).join(
        models.DayBook, models.DebitEntry.daybook_id == models.DayBook.id
    ).filter(func.lower(models.DebitEntry.name) == normalized_name).all()
    
    timeline = []
    
    for b in bills:
        rate_info = f" @ ₹{b.rate}/g" if b.is_rate_cut and b.rate else ""
        details = f"Purchase: {b.metal} {b.total_weight}g"
        if b.is_rate_cut:
            details += f" (Rate Cut{rate_info})"
        else:
            details += " (Pure metal)"
            
        timeline.append({
            "id": f"bill_{b.id}",
            "date": b.bill_date,
            "type": "bill",
            "reference": b.bill_no,
            "details": details,
            "amount": b.invoice_total or 0.0,
            "pure_weight": b.pure_weight or 0.0,
            "metal": b.metal,
            "is_rate_cut": b.is_rate_cut == 1
        })
        
    for d, db_date in debits:
        rate_cut = parse_rate_cut_payment(d.particulars)
        is_rc = rate_cut is not None
        rc_metal = ""
        rc_rate = 0.0
        rc_wt = 0.0
        details = d.remarks or "Supplier Payment"
        
        if rate_cut:
            rc_metal, rc_rate, rc_wt = rate_cut
            details = f"Rate Cut: {rc_wt:.3f}g {rc_metal} @ ₹{rc_rate}/g. " + (d.remarks or "")
            
        # Clean particulars representation
        ref_text = d.particulars
        ref_text = re.sub(r"^\[(?:UPI|OTHER)\]\s*", "", ref_text)
        ref_text = re.sub(r"\[RATE_CUT:[^\]]+\]\s*", "", ref_text)
        
        timeline.append({
            "id": f"payment_{d.id}",
            "date": db_date,
            "type": "payment",
            "reference": ref_text.strip(),
            "details": details,
            "amount": d.amount,
            "pure_weight": rc_wt,
            "metal": rc_metal,
            "is_rate_cut": is_rc,
            "rate": rc_rate
        })
        
    # Sort chronological ascending (oldest to newest)
    timeline.sort(key=lambda x: x["date"])
    return timeline


@app.post("/api/purchase/parties/{party_name}/payment")
def record_party_payment(party_name: str, payment: schemas.SupplierPaymentCreate, db: Session = Depends(get_db)):
    # 1. Format the particulars prefix based on the mode
    prefix = ""
    if payment.payment_mode == "UPI":
        prefix = "[UPI] "
    elif payment.payment_mode == "OTHER":
        prefix = "[OTHER] "
        
    # 2. Add rate cut tag if applicable
    rc_tag = ""
    if payment.is_rate_cut and payment.rate and payment.metal:
        # Calculate weight: Amount / Rate
        weight = payment.amount / payment.rate
        weight = round(weight, 3)
        rc_tag = f"[RATE_CUT:{payment.metal.upper()}|{payment.rate}|{weight}] "
        
    particulars = f"{prefix}{rc_tag}Supplier Payment"
    
    # 3. Get or create the DayBook for the given date
    db_daybook = db.query(models.DayBook).filter(models.DayBook.date == payment.date).first()
    if not db_daybook:
        # Create daybook carrying forward opening balance from previous day
        prev = crud.get_previous_daybook(db, payment.date)
        p_cash  = prev.closing_cash  if prev else 0.0
        p_upi   = prev.closing_upi   if prev else 0.0
        p_other = prev.closing_other if prev else 0.0
        p_upi_details = prev.closing_upi_details if (prev and prev.closing_upi_details) else "{}"
        
        db_daybook = models.DayBook(
            date=payment.date,
            opening_cash=p_cash,
            opening_upi=p_upi,
            opening_other=p_other,
            closing_cash=p_cash,
            closing_upi=p_upi,
            closing_other=p_other,
            opening_upi_details=p_upi_details,
            closing_upi_details=p_upi_details
        )
        db.add(db_daybook)
        db.commit()
        db.refresh(db_daybook)
        
    # 4. Create the DebitEntry
    debit_entry_data = schemas.DebitEntryCreate(
        name=party_name.strip(),
        particulars=particulars,
        amount=payment.amount,
        remarks=payment.remarks
      )
    
    # 5. Save using the existing crud method (which auto-triggers daybook recalculation & cascade)
    db_debit = crud.create_debit_entry(db, debit_entry_data, db_daybook.id)
    
    rate_cut_desc = ""
    if payment.is_rate_cut and payment.rate and payment.metal:
        rate_cut_desc = f" (Rate Cut: {payment.amount / payment.rate:.3f}g {payment.metal} @ ₹{payment.rate}/g)"
    log_system_action(db, "SUPPLIER_PAYMENT", f"Paid ₹{payment.amount} to supplier '{party_name}' via {payment.payment_mode}{rate_cut_desc}")
    return db_debit


@app.put("/api/purchase/parties/{party_id}", response_model=schemas.PurchasePartyResponse)
def update_purchase_party(party_id: int, party_update: schemas.PurchasePartyCreate, db: Session = Depends(get_db)):
    from sqlalchemy import func
    db_party = db.query(models.PurchaseParty).filter(models.PurchaseParty.id == party_id).first()
    if not db_party:
        raise HTTPException(status_code=404, detail="Supplier not found")
        
    # Check if name is being changed and is unique
    new_name = party_update.name.strip()
    if new_name.lower() != db_party.name.lower():
        existing = db.query(models.PurchaseParty).filter(
            func.lower(models.PurchaseParty.name) == new_name.lower()
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Supplier with this name already exists")
        db_party.name = new_name
        
    db_party.phone = party_update.phone.strip() if party_update.phone else None
    db_party.gstin = party_update.gstin.strip().upper() if party_update.gstin else None
    db_party.address = party_update.address.strip() if party_update.address else None
    db_party.opening_balance_cash = party_update.opening_balance_cash
    db_party.opening_balance_gold = party_update.opening_balance_gold
    db_party.opening_balance_silver = party_update.opening_balance_silver
    
    db.commit()
    db.refresh(db_party)
    log_system_action(db, "PARTY_UPDATE", f"Updated details for Supplier: {db_party.name} | Phone: {db_party.phone} | Balances: cash ₹{db_party.opening_balance_cash}, gold {db_party.opening_balance_gold}g, silver {db_party.opening_balance_silver}g")
    return db_party


@app.delete("/api/purchase/parties/{party_id}")
def delete_purchase_party(party_id: int, db: Session = Depends(get_db)):
    db_party = db.query(models.PurchaseParty).filter(models.PurchaseParty.id == party_id).first()
    if not db_party:
        raise HTTPException(status_code=404, detail="Supplier not found")
    name = db_party.name
    db.delete(db_party)
    db.commit()
    log_system_action(db, "PARTY_DELETE", f"Deleted Supplier: {name}")
    return {"message": "Supplier deleted successfully"}


@app.get("/api/system/logs", response_model=List[schemas.SystemLogResponse])
def get_system_logs(password: str, db: Session = Depends(get_db)):
    if password != "pooja123":
        raise HTTPException(status_code=401, detail="Unauthorized admin password required")
    return db.query(models.SystemLog).order_by(models.SystemLog.timestamp.desc()).all()


@app.get("/api/system/db/download")
def download_db_file(password: str):
    if password != "pooja123":
        raise HTTPException(status_code=401, detail="Unauthorized admin password required")
    import os
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, "..", "database.db")
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="Database file not found")
    return FileResponse(db_path, media_type="application/x-sqlite3", filename="pooja_jewellers_database.db")


@app.post("/api/system/db/restore")
def restore_db_file(password: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if password != "pooja123":
        raise HTTPException(status_code=401, detail="Unauthorized admin password required")
    
    if not file.filename.endswith(".db") and not file.filename.endswith(".sqlite"):
        raise HTTPException(status_code=400, detail="Invalid file type. Must be a .db SQLite database file.")
        
    import os
    import shutil
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, "..", "database.db")
    temp_path = os.path.join(base_dir, "..", "database_temp.db")
    
    try:
        # Save upload to temp
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # Safely close database connection pool
        engine.dispose()
        
        # Replace live database.db with uploaded file
        shutil.copy2(temp_path, db_path)
        
        if os.path.exists(temp_path):
            os.remove(temp_path)
            
        log_system_action(db, "SYSTEM_RESTORE", "System database file was successfully restored from an uploaded backup.")
        return {"success": True, "message": "Database file restored successfully."}
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise HTTPException(status_code=500, detail=f"Database restore failed: {e}")


@app.get("/api/system/network-time")
def get_network_time():
    """
    Fetches real-time network Indian Standard Time (IST / Asia/Kolkata) from internet APIs
    or HTTP Date headers, updating the backend global time offset.
    """
    global GLOBAL_TIME_OFFSET_SECONDS
    import urllib.request
    import json
    from email.utils import parsedate_to_datetime
    from datetime import datetime, timezone, timedelta

    ist_tz = timezone(timedelta(hours=5, minutes=30))

    # 1. High-reliability HTTP Date header strategy (Google / Cloudflare / GitHub)
    for head_url in ["https://www.google.com", "https://cloudflare.com", "https://api.github.com"]:
        try:
            req = urllib.request.Request(head_url, headers={"User-Agent": "Mozilla/5.0"}, method="HEAD")
            with urllib.request.urlopen(req, timeout=3) as resp:
                date_hdr = resp.headers.get("Date")
                if date_hdr:
                    gmt_dt = parsedate_to_datetime(date_hdr)
                    ist_dt = gmt_dt.astimezone(ist_tz)
                    date_str = ist_dt.strftime("%Y-%m-%d")
                    time_str = ist_dt.strftime("%H:%M:%S")
                    real_ts = ist_dt.timestamp()
                    GLOBAL_TIME_OFFSET_SECONDS = real_ts - time.time()
                    return {
                        "source": f"http_header ({head_url})",
                        "date": date_str,
                        "time": time_str,
                        "iso": ist_dt.isoformat(),
                        "timestamp": int(real_ts * 1000)
                    }
        except Exception:
            pass

    # 2. Secondary Internet attempt: worldtimeapi.org
    try:
        req = urllib.request.Request(
            "https://worldtimeapi.org/api/timezone/Asia/Kolkata",
            headers={"User-Agent": "Mozilla/5.0"}
        )
        with urllib.request.urlopen(req, timeout=3) as resp:
            if resp.status == 200:
                data = json.loads(resp.read().decode())
                dt_str = data.get("datetime", "")
                date_str = dt_str[:10]
                time_str = dt_str[11:19]
                real_ts = float(data.get("unixtime", time.time()))
                GLOBAL_TIME_OFFSET_SECONDS = real_ts - time.time()
                return {
                    "source": "worldtimeapi.org",
                    "date": date_str,
                    "time": time_str,
                    "iso": dt_str,
                    "timestamp": int(real_ts * 1000)
                }
    except Exception:
        pass

    # 3. Tertiary Internet attempt: timeapi.io
    try:
        req = urllib.request.Request(
            "https://timeapi.io/api/v1/time/current/zone?timeZone=Asia/Kolkata",
            headers={"User-Agent": "Mozilla/5.0"}
        )
        with urllib.request.urlopen(req, timeout=3) as resp:
            if resp.status == 200:
                data = json.loads(resp.read().decode())
                year = data.get("year")
                month = data.get("month")
                day = data.get("day")
                hour = data.get("hour")
                minute = data.get("minute")
                seconds = data.get("seconds")
                date_str = f"{year:04d}-{month:02d}-{day:02d}"
                time_str = f"{hour:02d}:{minute:02d}:{seconds:02d}"
                dt = datetime(year, month, day, hour, minute, seconds, tzinfo=ist_tz)
                real_ts = dt.timestamp()
                GLOBAL_TIME_OFFSET_SECONDS = real_ts - time.time()
                return {
                    "source": "timeapi.io",
                    "date": date_str,
                    "time": time_str,
                    "iso": f"{date_str}T{time_str}+05:30",
                    "timestamp": int(real_ts * 1000)
                }
    except Exception:
        pass

    # Fallback to server IST (with offset if known)
    now_ist = get_real_ist_now()
    date_str = now_ist.strftime("%Y-%m-%d")
    time_str = now_ist.strftime("%H:%M:%S")
    return {
        "source": "server_ist" if GLOBAL_TIME_OFFSET_SECONDS == 0 else "server_ist_offset",
        "date": date_str,
        "time": time_str,
        "iso": now_ist.isoformat(),
        "timestamp": int(now_ist.timestamp() * 1000)
    }


class TimeOffsetRequest(BaseModel):
    offset_seconds: float

@app.post("/api/system/set-time-offset")
def set_time_offset(body: TimeOffsetRequest):
    global GLOBAL_TIME_OFFSET_SECONDS
    GLOBAL_TIME_OFFSET_SECONDS = body.offset_seconds
    now_ist = get_real_ist_now()
    return {
        "status": "ok",
        "offset_seconds": GLOBAL_TIME_OFFSET_SECONDS,
        "current_ist": now_ist.isoformat(),
        "date": now_ist.strftime("%Y-%m-%d"),
        "time": now_ist.strftime("%H:%M:%S")
    }


# ─── BACKUP & SYSTEM AUDIT LOG ENDPOINTS ───

def log_system_event(db: Session, action: str, details: str, module: str = "GENERAL", user_name: str = "admin"):
    try:
        from datetime import datetime
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = models.SystemLog(
            timestamp=now_str,
            action=action,
            details=details,
            module=module,
            user_name=user_name
        )
        db.add(log_entry)
        db.commit()
    except Exception as e:
        print(f"[Log Error] Failed to log event: {e}")

@app.get("/api/backup/download")
def download_database_backup(db: Session = Depends(get_db)):
    import os
    from datetime import datetime
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, "..", "database.db")
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="Database file not found")
    
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"daybook_backup_{timestamp}.db"
    log_system_event(db, "BACKUP_DOWNLOAD", f"User downloaded database backup file: {filename}", module="BACKUP")
    return FileResponse(db_path, filename=filename, media_type="application/x-sqlite3")

@app.post("/api/backup/create")
def create_manual_backup(db: Session = Depends(get_db)):
    import os, shutil
    from datetime import datetime
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, "..", "database.db")
    backups_dir = os.path.join(base_dir, "..", "backups")
    if not os.path.exists(backups_dir):
        os.makedirs(backups_dir)
    
    if not os.path.exists(db_path):
        raise HTTPException(status_code=404, detail="Database file not found")
    
    now = datetime.now()
    timestamp_str = now.strftime("%Y-%m-%d_%H%M%S")
    backup_filename = f"db_manual_backup_{timestamp_str}.db"
    backup_file_path = os.path.join(backups_dir, backup_filename)
    
    shutil.copy2(db_path, backup_file_path)
    size_bytes = os.path.getsize(backup_file_path)
    
    log_system_event(db, "BACKUP_CREATE", f"Manual snapshot created: {backup_filename} ({round(size_bytes/1024, 1)} KB)", module="BACKUP")
    
    return {
        "success": True,
        "filename": backup_filename,
        "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
        "size_bytes": size_bytes,
        "message": f"Snapshot created: {backup_filename}"
    }

@app.get("/api/backup/list")
def list_backups():
    import os
    from datetime import datetime
    base_dir = os.path.dirname(os.path.abspath(__file__))
    backups_dir = os.path.join(base_dir, "..", "backups")
    if not os.path.exists(backups_dir):
        return []
    
    files = []
    for f in os.listdir(backups_dir):
        if f.endswith(".db"):
            fpath = os.path.join(backups_dir, f)
            mtime = os.path.getmtime(fpath)
            size = os.path.getsize(fpath)
            dt_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
            files.append({
                "filename": f,
                "timestamp": dt_str,
                "size_bytes": size,
                "mtime": mtime
            })
    files.sort(key=lambda x: x["mtime"], reverse=True)
    return files

@app.post("/api/backup/restore")
async def restore_database_backup(file: UploadFile = File(...), db: Session = Depends(get_db)):
    import os, shutil
    from datetime import datetime
    if not file.filename.endswith(".db"):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload a valid .db SQLite database file.")
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(base_dir, "..", "database.db")
    backups_dir = os.path.join(base_dir, "..", "backups")
    if not os.path.exists(backups_dir):
        os.makedirs(backups_dir)
    
    # 1. Create safety pre-restore backup
    safety_filename = f"pre_restore_safety_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.db"
    safety_path = os.path.join(backups_dir, safety_filename)
    if os.path.exists(db_path):
        shutil.copy2(db_path, safety_path)
    
    # 2. Save uploaded backup as temporary file
    temp_path = os.path.join(backups_dir, "temp_restore.db")
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # 3. Replace active DB
    db.close()
    shutil.copy2(temp_path, db_path)
    if os.path.exists(temp_path):
        os.remove(temp_path)
    
    log_system_event(db, "BACKUP_RESTORE", f"Database restored from uploaded file: {file.filename}", module="BACKUP")
    return {
        "success": True,
        "message": f"Database successfully restored from {file.filename}! Safety snapshot saved as {safety_filename}."
    }

@app.get("/api/system-logs")
def get_system_logs(
    limit: int = 100,
    module: Optional[str] = None,
    action: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.SystemLog)
    if module and module != "ALL":
        query = query.filter(models.SystemLog.module == module)
    if action and action != "ALL":
        query = query.filter(models.SystemLog.action == action)
    if search:
        s = f"%{search}%"
        query = query.filter(
            (models.SystemLog.details.like(s)) |
            (models.SystemLog.action.like(s))
        )
    query = query.order_by(models.SystemLog.id.desc()).limit(limit)
    return query.all()








